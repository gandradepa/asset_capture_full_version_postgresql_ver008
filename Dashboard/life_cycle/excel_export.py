#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Styled XLSX export for the Life Cycle Assessment dashboard.

Mirrors the UBC review-dashboard workbook styling used by the EL Review export
(navy #002145 header, slate borders, title + italic subtitle banner, hidden
gridlines, corner logo, a Summary sheet). A missing completeness value is
shaded salmon (#feb2a0), matching the on-screen table.
"""

import os
from io import BytesIO
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .completeness import COMPLETENESS_COLUMNS

# Palette (ARGB) — same anchors as the EL review export.
HEADER_FILL = "FF002145"
HEADER_TEXT = "FFFFFFFF"
BORDER_COLOR = "FFE2E8F0"
TITLE_COLOR = "FF002145"
SUBTITLE_COLOR = "FF64748B"
BAND_FILL = "FFEFF6FF"
MISSING_FILL = "FFFEB2A0"

# Classification cell fills / fonts mirror the dashboard chips.
CLASS_FILL = {"G": "FFDCFCE7", "Y": "FFFEF9C3", "R": "FFFEE2E2"}
CLASS_FONT = {"G": "FF166534", "Y": "FF854D0E", "R": "FF991B1B"}
CLASS_LABEL = {"G": "Good", "Y": "Caution", "R": "Critical"}

# (header label, row key, kind). kinds: text | number | class
COLUMNS = [
    ("QR Code", "QR Code", "text"),
    ("Description", "Description", "text"),
    ("Asset Tag", "Asset Tag", "text"),
    ("Building", "Property Name", "text"),
    ("Space Number", "Space Number", "text"),
    ("Space Name", "Space Name", "text"),
    ("Floor Name", "Floor Name", "text"),
    ("Make", "Make", "text"),
    ("Model", "Model", "text"),
    ("Serial Number", "Serial Number", "text"),
    ("Installation Date", "Installation Date", "text"),
    ("Years", "years", "number"),
    ("Classification", "classification", "class"),
]

# Completeness criterion — a blank Installation Date gets the salmon shade.
KEY_COLUMNS = COMPLETENESS_COLUMNS


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _cell_text(key: str, kind: str, value) -> str:
    if kind == "class":
        return CLASS_LABEL.get(value, "")
    if kind == "number":
        if _is_blank(value):
            return ""
        try:
            return f"{float(value):.2f}"
        except (TypeError, ValueError):
            return str(value)
    return "" if _is_blank(value) else str(value)


def _embed_corner_logo(ws, logo_path: Optional[str], anchor: str = "A1") -> bool:
    """Drop the UBC Facilities logo in the corner. Decorative — never fail the
    export over it (missing path / Pillow / IO error → skipped)."""
    if not logo_path:
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage  # needs Pillow
        if not os.path.isfile(logo_path):
            return False
        img = XLImage(logo_path)
        img.width = 70
        img.height = 60
        ws.add_image(img, anchor)
        return True
    except Exception as e:  # pragma: no cover - decorative
        print(f"[WARN] life_cycle excel_export logo embedding skipped: {e}")
        return False


def build_workbook(group_name: str, view_label: str, rows: list,
                   logo_path: Optional[str] = None) -> bytes:
    """Return XLSX bytes for the given (already filtered & ordered) rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = (view_label or "Life Cycle")[:31]
    ws.sheet_view.showGridLines = False

    logo_embedded = _embed_corner_logo(ws, logo_path, anchor="A1")
    title_col = 2 if logo_embedded else 1
    if logo_embedded:
        ws.column_dimensions["A"].width = 11

    title_cell = ws.cell(row=1, column=title_col,
                         value="UBC Facilities — Life Cycle Assessment Export")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=TITLE_COLOR)
    title_cell.alignment = Alignment(vertical="center", horizontal="left")

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    sub_text = (
        f"Asset Group: {group_name or '(all)'}   |   "
        f"View: {view_label}   |   "
        f"Generated {generated_at}   |   "
        f"{len(rows)} rows"
    )
    sub_cell = ws.cell(row=2, column=title_col, value=sub_text)
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color=SUBTITLE_COLOR)
    sub_cell.alignment = Alignment(vertical="center", horizontal="left")

    ws.row_dimensions[1].height = 36 if logo_embedded else 22
    ws.row_dimensions[2].height = 26 if logo_embedded else 18
    ws.row_dimensions[3].height = 6

    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_border = Border(
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
    )

    # Content-aware column widths (clamped).
    widths = []
    for header, key, kind in COLUMNS:
        longest = len(header)
        for row in rows:
            longest = max(longest, len(_cell_text(key, kind, row.get(key))))
        widths.append(max(12, min(42, longest + 2)))

    HEADER_ROW = 4
    for col_idx, (header, _key, _kind) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border
        ws.column_dimensions[c.column_letter].width = widths[col_idx - 1]
    ws.row_dimensions[HEADER_ROW].height = 26

    body_left = Alignment(horizontal="left", vertical="center")
    body_center = Alignment(horizontal="center", vertical="center")
    missing_fill = PatternFill(start_color=MISSING_FILL, end_color=MISSING_FILL, fill_type="solid")

    for offset, row in enumerate(rows):
        r = HEADER_ROW + 1 + offset
        for col_idx, (_header, key, kind) in enumerate(COLUMNS, start=1):
            raw = row.get(key)
            cell = ws.cell(row=r, column=col_idx)
            cell.border = cell_border

            if kind == "number":
                cell.alignment = body_center
                if not _is_blank(raw):
                    try:
                        cell.value = round(float(raw), 2)
                        cell.number_format = "0.00"
                    except (TypeError, ValueError):
                        cell.value = str(raw)
            elif kind == "class":
                cell.alignment = body_center
                code = str(raw or "").strip().upper()
                if code in CLASS_LABEL:
                    cell.value = CLASS_LABEL[code]
                    cell.fill = PatternFill(start_color=CLASS_FILL[code],
                                            end_color=CLASS_FILL[code], fill_type="solid")
                    cell.font = Font(name="Calibri", size=11, bold=True, color=CLASS_FONT[code])
            else:
                cell.alignment = body_left
                cell.value = "" if _is_blank(raw) else str(raw)

            # Shade blanks in the completeness criteria (matches the dashboard).
            if key in KEY_COLUMNS and _is_blank(raw):
                cell.fill = missing_fill

    _build_summary_sheet(wb, group_name, view_label, rows, logo_path)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(wb, group_name: str, view_label: str, rows: list,
                         logo_path: Optional[str] = None):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
    band_fill = PatternFill(start_color=BAND_FILL, end_color=BAND_FILL, fill_type="solid")
    label_font = Font(name="Calibri", size=11, bold=True, color=TITLE_COLOR)
    value_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=14, bold=True, color=TITLE_COLOR)

    logo_embedded = _embed_corner_logo(ws, logo_path, anchor="A1")
    title_col = 2 if logo_embedded else 1
    if logo_embedded:
        ws.column_dimensions["A"].width = 11
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 22

    ws.cell(row=1, column=title_col,
            value="UBC Facilities — Life Cycle Assessment Summary").font = title_font
    ws.cell(row=2, column=title_col,
            value=f"Generated {datetime.now():%Y-%m-%d %H:%M}").font = Font(
        name="Calibri", size=10, italic=True, color=SUBTITLE_COLOR)
    ws.column_dimensions["B" if logo_embedded else "A"].width = 30
    ws.column_dimensions["C" if logo_embedded else "B"].width = 26

    h1 = ws.cell(row=4, column=title_col, value="Metric")
    h2 = ws.cell(row=4, column=title_col + 1, value="Value")
    for c in (h1, h2):
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    tally = {"G": 0, "Y": 0, "R": 0}
    for r in rows:
        code = str(r.get("classification") or "").strip().upper()
        if code in tally:
            tally[code] += 1

    summary_rows = [
        ("Asset Group", group_name or "(all)"),
        ("View", view_label),
        ("Total rows", len(rows)),
        ("Good (under 8 years)", tally["G"]),
        ("Caution (8 to 10 years)", tally["Y"]),
        ("Critical (10 years or more)", tally["R"]),
    ]

    for idx, (label, value) in enumerate(summary_rows, start=5):
        a = ws.cell(row=idx, column=title_col, value=label)
        b = ws.cell(row=idx, column=title_col + 1, value=value)
        a.font = label_font
        b.font = value_font
        if idx % 2 == 0:
            a.fill = band_fill
            b.fill = band_fill
