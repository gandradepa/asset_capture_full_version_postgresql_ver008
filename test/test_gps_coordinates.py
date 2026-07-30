from __future__ import annotations

import pathlib
import sys
import unittest

from openpyxl import load_workbook


ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from API.validators_shared import format_gps_coordinates


class GpsCoordinateFormatterTests(unittest.TestCase):
    def test_valid_pair_is_comma_joined(self) -> None:
        self.assertEqual(format_gps_coordinates("49.2606051", "-123.2469569"), "49.260605,-123.246957")

    def test_missing_half_pair_is_blank(self) -> None:
        self.assertEqual(format_gps_coordinates("49.2606051", ""), "")
        self.assertEqual(format_gps_coordinates("", "-123.2469569"), "")

    def test_invalid_or_out_of_range_pair_is_blank(self) -> None:
        self.assertEqual(format_gps_coordinates("not-a-lat", "-123.2469569"), "")
        self.assertEqual(format_gps_coordinates("91", "-123.2469569"), "")
        self.assertEqual(format_gps_coordinates("49.2606051", "-181"), "")


class GpsCoordinateReviewSourceTests(unittest.TestCase):
    REVIEW_SOURCES = (
        ROOT / "review" / "Asset_dasboard_browser_ME" / "asset_plate_reviewer.py",
        ROOT / "review" / "Asset_dasboard_browser_BF" / "asset_plate_reviewer_bf.py",
        ROOT / "review" / "Asset_dashboard_browser_EL" / "Asset_dashboard_EL.py",
    )
    REVIEW_TEMPLATES = (
        ROOT / "review" / "Asset_dasboard_browser_ME" / "review_asset_templates" / "review.html",
        ROOT / "review" / "Asset_dasboard_browser_BF" / "review_asset_templates" / "review.html",
        ROOT / "review" / "Asset_dashboard_browser_EL" / "review_asset_templates" / "review.html",
    )

    def test_review_helpers_return_gps_coordinates(self) -> None:
        for source_path in self.REVIEW_SOURCES:
            text = source_path.read_text(encoding="utf-8")
            with self.subTest(path=str(source_path)):
                self.assertIn('"gps_coordinates": ""', text)
                self.assertIn('SELECT "GPS Coordinates (lat,long)" AS gps_coordinates', text)
                self.assertIn('result["gps_coordinates"]', text)
                self.assertIn("return qrdb.is_postgres() or os.path.exists(DB_PATH)", text)

    def test_review_templates_render_read_only_gps_field(self) -> None:
        for template_path in self.REVIEW_TEMPLATES:
            text = template_path.read_text(encoding="utf-8")
            with self.subTest(path=str(template_path)):
                self.assertIn("GPS Coordinates (lat,long)", text)
                self.assertIn("capture_info.gps_coordinates", text)
                self.assertNotIn('name="GPS Coordinates (lat,long)"', text)

    def test_capture_write_path_syncs_merged_column(self) -> None:
        source_text = (ROOT / "asset_capture_app_dev" / "app.py").read_text(encoding="utf-8")
        self.assertIn("format_gps_coordinates(latitude, longitude)", source_text)
        self.assertIn('"GPS Coordinates (lat,long)" if "GPS Coordinates (lat,long)" in cols else None', source_text)
        self.assertIn("_quote_ident(gps_col)", source_text)


class GpsCoordinateSdiExportTests(unittest.TestCase):
    def test_sdi_template_has_gps_coordinates_header(self) -> None:
        template_path = ROOT / "SDI_process" / "template" / "Import Assets-TEMPLATE-082923.xlsx"
        workbook = load_workbook(template_path, read_only=True, data_only=False)
        try:
            sheet = workbook["UsrAsset"]
            headers = [sheet.cell(row=9, column=idx).value for idx in range(1, sheet.max_column + 1)]
        finally:
            workbook.close()

        self.assertIn("GPS Coordinates (lat,long)", headers)

    def test_sdi_export_sources_gps_from_qr_codes(self) -> None:
        source_text = (ROOT / "SDI_process" / "app.py").read_text(encoding="utf-8")
        self.assertIn('GPS_COORDINATES_COL = "GPS Coordinates (lat,long)"', source_text)
        self.assertIn('GPS_COORDINATES_DB_ALIASES = (GPS_COORDINATES_COL, "GPS Coordinates (lat, long)")', source_text)
        self.assertIn('GPS_COORDINATES_COL, "Diameter"', source_text)
        self.assertIn('AS "{GPS_COORDINATES_COL}"', source_text)
        self.assertIn('_merge_qr_gps_coordinates(df, conn, overwrite_from_db=True)', source_text)
        self.assertIn("2026-06-23_sdi_package_gps_coordinates.sql", source_text)
        self.assertIn('print(f"[ERROR] in build_unpackaged_dataset (merging data): {repr(e)}")\n            return pd.DataFrame()', source_text)
        self.assertIn("def _build_qr_process_lookup", source_text)
        self.assertIn('df = df[pd.to_numeric(df["max_col_process"], errors="coerce").fillna(-1).astype(int) == 0]', source_text)


if __name__ == "__main__":
    unittest.main()
