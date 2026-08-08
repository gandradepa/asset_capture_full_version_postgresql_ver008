## /home/developer/SDI_process/app.py

import os
import re
import sys
import sqlite3
import glob
import json
import logging
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List

import pandas as pd
import db as qrdb  # backend-agnostic QR_codes DB layer (aliased: 'db' here is Flask-SQLAlchemy from auth_model)
from flask import (
    Flask, render_template, redirect, url_for, flash,
    request, send_file, Blueprint, jsonify, g
)
from openpyxl import load_workbook

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.normpath(os.path.join(THIS_DIR, ".."))
if REPO_ROOT not in sys.path:
    sys.path.append(REPO_ROOT)

try:
    from electrical_equipment_rules import (
        derive_electrical_equipment_type,
        parse_electrical_equipment_metadata,
    )
except ModuleNotFoundError:
    _ELECTRICAL_EQUIPMENT_TYPE_MAP = {
        "MDP": "Main Distribution Panel",
        "CDP": "Central Distribution Panel",
        "SPL": "Splitter",
        "MCC": "Motor Control Center",
        "PNL": "Panel",
        "SWBD": "Switchboard",
        "ATS": "Automatic Transfer Switch",
        "TX": "Transformer",
    }
    _ELECTRICAL_POWER_TYPE_CODES = {"NES", "NE", "NS", "ES", "N", "E", "S"}
    _ELECTRICAL_POWER_TYPE_RE = re.compile(r"NES|NE|NS|ES|N|E|S")

    def derive_electrical_equipment_type(tag_value: object) -> str:
        tag_upper = str(tag_value or "").strip().upper()
        if not tag_upper:
            return ""
        for code, name in sorted(_ELECTRICAL_EQUIPMENT_TYPE_MAP.items(), key=lambda item: len(item[0]), reverse=True):
            if tag_upper.startswith(code):
                return name
        return ""

    def parse_electrical_equipment_metadata(tag_value: object) -> tuple[str, str]:
        tag_upper = str(tag_value or "").strip().upper()
        if not tag_upper:
            return "", ""
        power_type = ""
        parts = tag_upper.split("-")
        if len(parts) > 1:
            match = _ELECTRICAL_POWER_TYPE_RE.search(parts[1])
            if match and match.group(0) in _ELECTRICAL_POWER_TYPE_CODES:
                power_type = match.group(0)
        return power_type, derive_electrical_equipment_type(tag_upper)

## Import authentication and environment variable libraries
from flask_login import login_user, logout_user, login_required, current_user
from dotenv import load_dotenv

## Add the shared auth_service directory to Python's path
sys.path.append('/home/developer/auth_service')
from auth_model import db, bcrypt, User, ensure_user_access_table, has_permission, is_admin, require_permission, access_denied_response
from auth_controller import login_manager
from audit.logger import log_change as _audit_log_change

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "template")
STATIC_DIR = os.path.join(BASE_DIR, "static")
DB_PATH = r"/home/developer/asset_capture_app_dev/data/QR_codes.db"
TEMPLATE_PATH = r"/home/developer/SDI_process/template/Import Assets-TEMPLATE-082923.xlsx"
JSON_OUTPUT_DIR = os.path.abspath(os.getenv("JSON_OUTPUT_DIR") or os.path.join(BASE_DIR, "sdi_json_output"))
REVIEW_JSON_DIR = os.path.abspath(os.getenv("JSON_DIR") or "/home/developer/Output_jason_api")
REVIEW_JSON_PROCESSED_LOGS = {
    "ME": os.path.join(os.path.dirname(DB_PATH), "processed_json.log"),
    "BF": os.path.join(os.path.dirname(DB_PATH), "processed_json_bf.log"),
    "EL": os.path.join(os.path.dirname(DB_PATH), "processed_json_el.log"),
}

LOGO_MAIN_NAME = "ubc_logo.jpg"
LOGO_FAC_NAME = "ubc-facilities_logo.jpg"
INVALID_QR_KEY_VALUES = {"", "none", "nan", "null"}

# -----------------------------------------------------------------------------
# Flask App Configuration
# -----------------------------------------------------------------------------
load_dotenv('/home/developer/auth_service.env', override=True)

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR, static_url_path="/static")

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URI')
app.config['SESSION_COOKIE_DOMAIN']    = os.getenv('SESSION_COOKIE_DOMAIN')
app.config['SESSION_COOKIE_SAMESITE']  = 'None'
app.config['SESSION_COOKIE_SECURE']    = True
app.config['SESSION_COOKIE_HTTPONLY']  = True
app.config['REMEMBER_COOKIE_SAMESITE'] = 'None'
app.config['REMEMBER_COOKIE_SECURE']   = True
app.config['REMEMBER_COOKIE_HTTPONLY'] = True

# Staging/local QA over plain HTTP (e.g. an SSH-tunnelled PG staging instance): relax the
# cross-subdomain Secure cookies so login works without HTTPS. Gated by an env var, so this
# is a NO-OP in production (the var is unset there) and changes nothing on the live service.
if os.getenv('STAGING_INSECURE_COOKIES') == '1':
    app.config['SESSION_COOKIE_DOMAIN']  = None
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['SESSION_COOKIE_SECURE']   = False
    app.config['REMEMBER_COOKIE_SAMESITE'] = 'Lax'
    app.config['REMEMBER_COOKIE_SECURE']   = False

db.init_app(app)
bcrypt.init_app(app)
login_manager.init_app(app)

# -----------------------------------------------------------------------------
# Columns & Mappings
# -----------------------------------------------------------------------------
GPS_COORDINATES_COL = "GPS Coordinates (lat,long)"
GPS_COORDINATES_DB_ALIASES = (GPS_COORDINATES_COL, "GPS Coordinates (lat, long)")

MASTER_COLS = [
    "id_print_out", "QR Code", "Building", "Description", "Asset Group", "UBC Tag", "Equipment ID", "Equipment Type", "Serial", "Model",
    "Manufacturer", "Attribute", "Amperage Rating", "Supply From", "Volts", "Location", "Space", "Floor Code",
    GPS_COORDINATES_COL, "Diameter", "Technical Safety BC", "Year", "Installation Date"
]
PACKAGE_ONLY_COLS = [
    "Voltage Rating",
    "Voltage Rating (UoM)",
    "Amperage Rating (UoM)",
    "Fed From Equipment ID",
    "Fed From Amperage Rating",
    "Fed From Amperage Rating (UoM)",
    "Power Type",
    "Power Rating",
    "Power Rating (UoM)",
]
PRINT_OUT_COLS = MASTER_COLS + PACKAGE_ONLY_COLS + ["print_out", "date", "time"]
COLUMN_RENAME_MAP: Dict[str, str] = {
    "QR Code": "Code", "Building": "Property", "Description": "Description",
    "Asset Group": "Asset Group", "UBC Tag": "Asset Tag", "Serial": "Serial Number",
    "Model": "Model", "Manufacturer": "Make", "Attribute": "Attribute Set",
    "Location": "Space Details", "Space": "Space.Space number",
    "Floor Code": "Space.Floor.Floor", "Diameter": "Diameter", "Technical Safety BC": "Previous (OLD) ID",
    "Year": "Date Of Manufacture Or Construction", "Installation Date": "Installation Date",
    GPS_COORDINATES_COL: GPS_COORDINATES_COL,
}
CONST_COLS: Dict[str, object] = {
    "Is Missing (Y/N)": False, "Simple": True, "Is Planned Maintenance Required? (Y/N)": False,
    "Status.System name": "UsrBaseAssetInUse", "Status.Business object": "UsrAsset",
}

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def table_exists(conn, table_name):
    return qrdb.has_table(conn, table_name)

def _read_sql(sql, conn, **kwargs):
    # pandas needs a real DBAPI/SQLAlchemy handle, not our wrapper -> pass the raw conn.
    return pd.read_sql_query(sql, qrdb.raw_conn(conn), **kwargs)

def ensure_columns_and_order(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df = df.loc[:,~df.columns.duplicated()]
    for c in MASTER_COLS:
        if c not in df.columns:
            df[c] = ""
    return df[MASTER_COLS]

def filter_approved(df: pd.DataFrame) -> pd.DataFrame:
    if "Approved" not in df.columns:
        return df.copy()
    return df.loc[df["Approved"].astype(str) == "1"].copy()

def _normalize_qr_key_series(series: pd.Series) -> pd.Series:
    normalized = series.fillna("").astype(str).str.strip()
    invalid_mask = normalized.str.lower().isin(INVALID_QR_KEY_VALUES)
    return normalized.mask(invalid_mask, '').str.upper()

def _merge_qr_gps_coordinates(
    df: pd.DataFrame,
    conn: sqlite3.Connection,
    *,
    overwrite_from_db: bool = True,
) -> pd.DataFrame:
    result = df.copy()
    if GPS_COORDINATES_COL not in result.columns:
        result[GPS_COORDINATES_COL] = ""
    if result.empty or "QR Code" not in result.columns or not table_exists(conn, "QR_codes"):
        return result

    qr_columns = set(qrdb.table_columns(conn, "QR_codes"))
    source_gps_col = next((col for col in GPS_COORDINATES_DB_ALIASES if col in qr_columns), None)
    if "QR_code_ID" not in qr_columns or source_gps_col is None:
        return result

    gps_lookup = _read_sql(
        f'''
        SELECT
            "QR_code_ID",
            "{source_gps_col}" AS "{GPS_COORDINATES_COL}"
        FROM "QR_codes"
        ''',
        conn,
    )
    if gps_lookup.empty:
        return result

    gps_lookup["__qr_gps_join_key"] = _normalize_qr_key_series(gps_lookup["QR_code_ID"])
    gps_lookup[GPS_COORDINATES_COL] = gps_lookup[GPS_COORDINATES_COL].fillna("").astype(str).str.strip()
    gps_lookup = gps_lookup[
        (gps_lookup["__qr_gps_join_key"] != "")
        & (gps_lookup[GPS_COORDINATES_COL] != "")
    ].copy()
    if gps_lookup.empty:
        return result

    gps_lookup = gps_lookup.drop_duplicates(subset=["__qr_gps_join_key"], keep="first")
    gps_lookup = gps_lookup[["__qr_gps_join_key", GPS_COORDINATES_COL]].rename(
        columns={GPS_COORDINATES_COL: "__qr_gps_from_db"}
    )

    result["__qr_gps_join_key"] = _normalize_qr_key_series(result["QR Code"])
    result = pd.merge(result, gps_lookup, on="__qr_gps_join_key", how="left")

    current_gps = result[GPS_COORDINATES_COL].fillna("").astype(str).str.strip()
    database_gps = result["__qr_gps_from_db"].fillna("").astype(str).str.strip()
    if overwrite_from_db:
        result[GPS_COORDINATES_COL] = database_gps.where(database_gps != "", current_gps)
    else:
        result[GPS_COORDINATES_COL] = current_gps.where(current_gps != "", database_gps)

    return result.drop(columns=["__qr_gps_join_key", "__qr_gps_from_db"])

def _build_qr_codes_lookup(conn: sqlite3.Connection) -> pd.DataFrame:
    rowid_expr = "ctid" if qrdb.is_postgres() else "rowid"  # PG has no rowid; ctid is the physical-row tiebreaker
    qr_codes_df = _read_sql(
        f'''
        SELECT
            {rowid_expr} AS qr_rowid,
            "QR_code_ID",
            "Space",
            "Space Details",
            "Floor Code",
            "sdi",
            "date_set",
            "installation_date" AS "Installation Date"
        FROM "QR_codes"
        ''',
        conn,
    )
    if qr_codes_df.empty:
        return pd.DataFrame(columns=["join_key", "Space", "Space Details", "Floor Code", "sdi", "Installation Date"])

    qr_codes_df["join_key"] = _normalize_qr_key_series(qr_codes_df["QR_code_ID"])
    qr_codes_df = qr_codes_df[qr_codes_df["join_key"] != ""].copy()
    if qr_codes_df.empty:
        return pd.DataFrame(columns=["join_key", "Space", "Space Details", "Floor Code", "sdi", "Installation Date"])

    for col in ("Space", "Space Details", "Floor Code", "sdi", "Installation Date"):
        if col not in qr_codes_df.columns:
            qr_codes_df[col] = ""

    qr_codes_df["sdi_rank"] = (qr_codes_df["sdi"].fillna("").astype(str).str.strip() != "1").astype(int)
    qr_codes_df["detail_rank"] = (
        qr_codes_df["Space"].fillna("").astype(str).str.strip().ne("").astype(int)
        + qr_codes_df["Space Details"].fillna("").astype(str).str.strip().ne("").astype(int)
        + qr_codes_df["Floor Code"].fillna("").astype(str).str.strip().ne("").astype(int)
    )
    qr_codes_df["date_rank"] = pd.to_datetime(qr_codes_df["date_set"], errors="coerce")

    # Keep the strongest location row per QR code so one asset cannot fan out into duplicates.
    qr_codes_df = qr_codes_df.sort_values(
        by=["join_key", "sdi_rank", "detail_rank", "date_rank", "qr_rowid"],
        ascending=[True, False, False, False, False],
    ).drop_duplicates(subset=["join_key"], keep="first")

    return qr_codes_df[["join_key", "Space", "Space Details", "Floor Code", "sdi", "Installation Date"]]

def _build_qr_process_lookup(conn: sqlite3.Connection) -> pd.DataFrame:
    if not table_exists(conn, "QR_code_assets"):
        return pd.DataFrame(columns=["join_key", "max_col_process"])

    asset_columns = set(qrdb.table_columns(conn, "QR_code_assets"))
    if "code_assets" not in asset_columns or "Col_process" not in asset_columns:
        return pd.DataFrame(columns=["join_key", "max_col_process"])

    process_df = _read_sql(
        '''
        SELECT "code_assets", "Col_process"
        FROM "QR_code_assets"
        ''',
        conn,
    )
    if process_df.empty:
        return pd.DataFrame(columns=["join_key", "max_col_process"])

    process_df["join_key"] = _normalize_qr_key_series(
        process_df["code_assets"].fillna("").astype(str).str.split().str[0]
    )
    process_df["process_value"] = pd.to_numeric(process_df["Col_process"], errors="coerce")
    process_df = process_df[
        (process_df["join_key"] != "")
        & process_df["process_value"].notna()
    ].copy()
    if process_df.empty:
        return pd.DataFrame(columns=["join_key", "max_col_process"])

    return (
        process_df.groupby("join_key", as_index=False)["process_value"]
        .max()
        .rename(columns={"process_value": "max_col_process"})
    )

def ensure_qr_codes_integrity_guards(conn: sqlite3.Connection) -> None:
    if qrdb.is_postgres():
        # PG already enforces these via the PK on QR_code_ID + chk_qr_not_placeholder
        # (installed in the C2/C3 migration). The SQLite DDL below is SQLite-specific.
        return
    cur = conn.cursor()
    cur.execute(
        '''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_qr_codes_qr_code_id_norm_unique
        ON QR_codes (UPPER(TRIM(COALESCE(CAST(QR_code_ID AS TEXT), ''))))
        WHERE TRIM(COALESCE(CAST(QR_code_ID AS TEXT), '')) != ''
          AND LOWER(TRIM(COALESCE(CAST(QR_code_ID AS TEXT), ''))) NOT IN ('none', 'nan', 'null')
        '''
    )
    cur.execute(
        '''
        CREATE TRIGGER IF NOT EXISTS trg_qr_codes_reject_placeholder_insert
        BEFORE INSERT ON QR_codes
        FOR EACH ROW
        WHEN LOWER(TRIM(COALESCE(CAST(NEW.QR_code_ID AS TEXT), ''))) IN ('', 'none', 'nan', 'null')
        BEGIN
            SELECT RAISE(ABORT, 'Placeholder QR_code_ID values are not allowed');
        END;
        '''
    )
    cur.execute(
        '''
        CREATE TRIGGER IF NOT EXISTS trg_qr_codes_reject_placeholder_update
        BEFORE UPDATE OF QR_code_ID ON QR_codes
        FOR EACH ROW
        WHEN LOWER(TRIM(COALESCE(CAST(NEW.QR_code_ID AS TEXT), ''))) IN ('', 'none', 'nan', 'null')
        BEGIN
            SELECT RAISE(ABORT, 'Placeholder QR_code_ID values are not allowed');
        END;
        '''
    )

def _normalize_amperage_columns(df: pd.DataFrame, *, keep_legacy_ampere: bool = False) -> pd.DataFrame:
    if df.empty:
        return df

    normalized = df.copy()
    canonical = pd.Series("", index=normalized.index, dtype="object")
    if "Amperage Rating" in normalized.columns:
        canonical = normalized["Amperage Rating"].fillna("").astype(str).str.strip()
    if "Ampere" in normalized.columns:
        legacy = normalized["Ampere"].fillna("").astype(str).str.strip()
        canonical = canonical.where(canonical != "", legacy)
    canonical = canonical.str.extract(r"(\d{1,4})", expand=False).fillna("")
    uom = pd.Series("", index=normalized.index, dtype="object")
    uom = uom.mask(canonical != "", "AMP")
    normalized["Amperage Rating"] = canonical
    normalized["Amperage Rating (UoM)"] = uom
    if keep_legacy_ampere or "Ampere" in normalized.columns:
        normalized["Ampere"] = canonical
    return normalized

def _normalize_fed_from_amperage_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    normalized = df.copy()
    canonical = pd.Series("", index=normalized.index, dtype="object")
    if "Fed From Amperage Rating" in normalized.columns:
        canonical = (
            normalized["Fed From Amperage Rating"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.extract(r"(\d{1,4})", expand=False)
            .fillna("")
        )
    normalized["Fed From Amperage Rating"] = canonical
    fed_from_uom = pd.Series("", index=normalized.index, dtype="object")
    fed_from_uom = fed_from_uom.mask(canonical != "", "A")
    normalized["Fed From Amperage Rating (UoM)"] = fed_from_uom
    return normalized

def _normalize_fed_from_equipment_id_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    normalized = df.copy()
    canonical = pd.Series("", index=normalized.index, dtype="object")
    if "Fed From Equipment ID" in normalized.columns:
        canonical = normalized["Fed From Equipment ID"].fillna("").astype(str).str.strip()
    if "Supply From" in normalized.columns:
        supply_from = normalized["Supply From"].fillna("").astype(str).str.strip()
        canonical = canonical.where(canonical != "", supply_from)
        normalized["Supply From"] = supply_from.where(supply_from != "", canonical)
    normalized["Fed From Equipment ID"] = canonical
    return normalized

def _normalize_power_rating_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    normalized = df.copy()
    rating = pd.Series("", index=normalized.index, dtype="object")
    if "Power Rating" in normalized.columns:
        rating = normalized["Power Rating"].fillna("").astype(str).str.strip()
    uom = pd.Series("", index=normalized.index, dtype="object")
    if "Power Rating (UoM)" in normalized.columns:
        uom = normalized["Power Rating (UoM)"].fillna("").astype(str).str.strip()
    uom = uom.where(rating != "", '')
    normalized["Power Rating"] = rating
    normalized["Power Rating (UoM)"] = uom
    return normalized

# Canonical storage keeps the bare voltage value ("208/120"); the unit lives in
# "Voltage Rating (UoM)" as VLT. Legacy values may still embed the letters.
_VOLTAGE_UNIT_SUFFIX_RE = re.compile(r"(?<=\d)\s*V(?:AC|DC|OLTS?)?\b", re.IGNORECASE)

def _normalize_voltage_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    normalized = df.copy()
    rating = pd.Series("", index=normalized.index, dtype="object")
    if "Voltage Rating" in normalized.columns:
        rating = normalized["Voltage Rating"].fillna("").astype(str).str.strip()
    if "Volts" in normalized.columns:
        legacy = normalized["Volts"].fillna("").astype(str).str.strip()
        rating = rating.where(rating != "", legacy)
    rating = rating.str.replace(_VOLTAGE_UNIT_SUFFIX_RE, "", regex=True).str.strip()
    uom = pd.Series("", index=normalized.index, dtype="object")
    if "Voltage Rating (UoM)" in normalized.columns:
        uom = normalized["Voltage Rating (UoM)"].fillna("").astype(str).str.strip()
    uom = uom.where(rating != "", '')
    uom = uom.mask((rating != "") & (uom == ""), "VLT")
    normalized["Voltage Rating"] = rating
    normalized["Voltage Rating (UoM)"] = uom
    if "Volts" in normalized.columns:
        normalized["Volts"] = rating
    return normalized

def _normalize_equipment_id_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    normalized = df.copy()
    equipment = pd.Series("", index=normalized.index, dtype="object")
    if "Equipment ID" in normalized.columns:
        equipment = normalized["Equipment ID"].fillna("").astype(str).str.strip()

    fallback = pd.Series("", index=normalized.index, dtype="object")
    if "UBC Tag" in normalized.columns:
        fallback = normalized["UBC Tag"].fillna("").astype(str).str.strip()
    elif "UBC Asset Tag" in normalized.columns:
        fallback = normalized["UBC Asset Tag"].fillna("").astype(str).str.strip()

    if "Attribute" in normalized.columns:
        electrical_mask = normalized["Attribute"].fillna("").astype(str).str.strip().str.casefold() == "electrical"
    else:
        electrical_mask = fallback != ""

    normalized["Equipment ID"] = equipment
    needs_fill = electrical_mask & (normalized["Equipment ID"] == "") & (fallback != "")
    normalized.loc[needs_fill, "Equipment ID"] = fallback.loc[needs_fill]
    return normalized

def _normalize_equipment_type_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    normalized = df.copy()
    equipment_type = pd.Series("", index=normalized.index, dtype="object")
    if "Equipment Type" in normalized.columns:
        equipment_type = normalized["Equipment Type"].fillna("").astype(str).str.strip()

    source = pd.Series("", index=normalized.index, dtype="object")
    if "Equipment ID" in normalized.columns:
        source = normalized["Equipment ID"].fillna("").astype(str).str.strip()

    fallback = pd.Series("", index=normalized.index, dtype="object")
    if "UBC Tag" in normalized.columns:
        fallback = normalized["UBC Tag"].fillna("").astype(str).str.strip()
    elif "UBC Asset Tag" in normalized.columns:
        fallback = normalized["UBC Asset Tag"].fillna("").astype(str).str.strip()
    source = source.where(source != "", fallback)

    if "Attribute" in normalized.columns:
        electrical_mask = normalized["Attribute"].fillna("").astype(str).str.strip().str.casefold() == "electrical"
    else:
        electrical_mask = source != ""

    normalized["Equipment Type"] = equipment_type
    derived_type = source.apply(derive_electrical_equipment_type)
    needs_fill = electrical_mask & (normalized["Equipment Type"] == "") & (derived_type != "")
    normalized.loc[needs_fill, "Equipment Type"] = derived_type.loc[needs_fill]
    return normalized

def _normalize_export_source_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized = _normalize_voltage_columns(df)
    normalized = _normalize_amperage_columns(normalized)
    normalized = _normalize_fed_from_equipment_id_column(normalized)
    normalized = _normalize_fed_from_amperage_column(normalized)
    normalized = _normalize_power_rating_columns(normalized)
    normalized = _normalize_equipment_id_column(normalized)
    normalized = _normalize_equipment_type_column(normalized)
    if "Installation Date" in normalized.columns:
        parsed_dates = pd.to_datetime(
            normalized["Installation Date"].fillna("").astype(str).str.strip(),
            format="%Y-%m-%d", errors="coerce",
        )
        normalized["Installation Date"] = parsed_dates.dt.strftime("%Y-%m-%d").fillna("")
    return normalized

def _ensure_package_amperage_columns(conn: sqlite3.Connection, table_name: str) -> None:
    cur = conn.cursor()
    existing_cols = set(qrdb.table_columns(conn, table_name))
    if not existing_cols:
        return
    if "Installation Date" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Installation Date" TEXT')
        existing_cols.add("Installation Date")
    if GPS_COORDINATES_COL not in existing_cols:
        if qrdb.is_postgres():
            raise RuntimeError(
                f'Missing "{GPS_COORDINATES_COL}" on {table_name}. '
                "Run scripts/migrations/2026-06-23_sdi_package_gps_coordinates.sql "
                "as the PostgreSQL table owner before creating SDI packages."
            )
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{GPS_COORDINATES_COL}" TEXT')
        existing_cols.add(GPS_COORDINATES_COL)
    if "Equipment ID" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Equipment ID" TEXT')
        existing_cols.add("Equipment ID")
    if "Equipment Type" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Equipment Type" TEXT')
        existing_cols.add("Equipment Type")
    if "Power Type" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Power Type" TEXT')
        existing_cols.add("Power Type")
    if "Voltage Rating" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Voltage Rating" TEXT')
        existing_cols.add("Voltage Rating")
    if "Voltage Rating (UoM)" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Voltage Rating (UoM)" TEXT')
        existing_cols.add("Voltage Rating (UoM)")
    if "Fed From Equipment ID" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Fed From Equipment ID" TEXT')
        existing_cols.add("Fed From Equipment ID")
    tag_source_col = "UBC Tag" if "UBC Tag" in existing_cols else None
    if tag_source_col:
        electrical_filter = ""
        if "Attribute" in existing_cols:
            electrical_filter = ' AND LOWER(TRIM(COALESCE("Attribute", \'\'))) = \'electrical\''
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Equipment ID" = TRIM(COALESCE("{tag_source_col}", ''))
             WHERE TRIM(COALESCE("Equipment ID", '')) = ''
               AND TRIM(COALESCE("{tag_source_col}", '')) != ''
               {electrical_filter}
            '''
        )
        attribute_expr = '"Attribute"' if "Attribute" in existing_cols else "''"
        rowid_expr = "ctid" if qrdb.is_postgres() else "rowid"   # PG has no rowid
        rows = cur.execute(
            f'''
            SELECT
                {rowid_expr} AS _rowid,
                "Equipment Type",
                "Equipment ID",
                "{tag_source_col}" AS "Tag Source",
                {attribute_expr} AS "Attribute"
            FROM "{table_name}"
            '''
        ).fetchall()
        row_match = "ctid = ?::tid" if qrdb.is_postgres() else "rowid = ?"
        for rowid, current_type, equipment_id, tag_source, attribute_value in rows:
            if "Attribute" in existing_cols and str(attribute_value or "").strip().casefold() != "electrical":
                continue
            source_value = str(equipment_id or "").strip() or str(tag_source or "").strip()
            desired_type = derive_electrical_equipment_type(source_value)
            if str(current_type or "").strip() or not desired_type:
                continue
            cur.execute(
                f'UPDATE "{table_name}" SET "Equipment Type" = ? WHERE {row_match}',
                (desired_type, rowid),
            )
    if "Volts" in existing_cols:
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Voltage Rating" = TRIM(COALESCE("Volts", ''))
             WHERE TRIM(COALESCE("Voltage Rating", '')) = ''
               AND TRIM(COALESCE("Volts", '')) != ''
            '''
        )
    if "Volts" in existing_cols:
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Volts" = TRIM(COALESCE("Voltage Rating", ''))
             WHERE TRIM(COALESCE("Volts", '')) = ''
               AND TRIM(COALESCE("Voltage Rating", '')) != ''
            '''
        )
    if "Supply From" in existing_cols:
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Fed From Equipment ID" = TRIM(COALESCE("Supply From", ''))
             WHERE TRIM(COALESCE("Fed From Equipment ID", '')) = ''
               AND TRIM(COALESCE("Supply From", '')) != ''
            '''
        )
    if (
        "QR Code" in existing_cols
        and "Building" in existing_cols
        and "Fed From Equipment ID" in existing_cols
        and table_exists(conn, "sdi_dataset_EL")
    ):
        electrical_filter = ""
        if "Attribute" in existing_cols:
            electrical_filter = ' AND LOWER(TRIM(COALESCE(child."Attribute", \'\'))) = \'electrical\''
        cur.execute(
            f'''
            UPDATE "{table_name}" AS child
               SET "Fed From Equipment ID" = COALESCE((
                       SELECT CASE
                           WHEN TRIM(COALESCE(parent."Fed From Equipment ID", '')) != '' THEN TRIM(COALESCE(parent."Fed From Equipment ID", ''))
                           ELSE TRIM(COALESCE(parent."Supply From", ''))
                       END
                       FROM "sdi_dataset_EL" AS parent
                       WHERE TRIM(COALESCE(parent."Building", '')) = TRIM(COALESCE(child."Building", ''))
                         AND UPPER(TRIM(COALESCE(parent."QR Code", ''))) = UPPER(TRIM(COALESCE(child."QR Code", '')))
                       LIMIT 1
                   ), TRIM(COALESCE(child."Supply From", '')))
             WHERE 1 = 1
               {electrical_filter}
            '''
        )
    if (
        "QR Code" in existing_cols
        and "Building" in existing_cols
        and "Voltage Rating" in existing_cols
        and "Voltage Rating (UoM)" in existing_cols
        and table_exists(conn, "sdi_dataset_EL")
    ):
        electrical_filter = ""
        if "Attribute" in existing_cols:
            electrical_filter = ' AND LOWER(TRIM(COALESCE(child."Attribute", \'\'))) = \'electrical\''
        cur.execute(
            f'''
            UPDATE "{table_name}" AS child
               SET "Voltage Rating" = COALESCE((
                       SELECT CASE
                           WHEN TRIM(COALESCE(parent."Voltage Rating", '')) != '' THEN TRIM(COALESCE(parent."Voltage Rating", ''))
                           ELSE TRIM(COALESCE(parent."Volts", ''))
                       END
                       FROM "sdi_dataset_EL" AS parent
                       WHERE TRIM(COALESCE(parent."Building", '')) = TRIM(COALESCE(child."Building", ''))
                         AND UPPER(TRIM(COALESCE(parent."QR Code", ''))) = UPPER(TRIM(COALESCE(child."QR Code", '')))
                       LIMIT 1
                   ), ''),
                   "Voltage Rating (UoM)" = COALESCE((
                       SELECT CASE
                           WHEN TRIM(COALESCE(parent."Voltage Rating", '')) != '' THEN TRIM(COALESCE(parent."Voltage Rating (UoM)", ''))
                           WHEN TRIM(COALESCE(parent."Volts", '')) != '' THEN 'VLT'
                           ELSE ''
                       END
                       FROM "sdi_dataset_EL" AS parent
                       WHERE TRIM(COALESCE(parent."Building", '')) = TRIM(COALESCE(child."Building", ''))
                         AND UPPER(TRIM(COALESCE(parent."QR Code", ''))) = UPPER(TRIM(COALESCE(child."QR Code", '')))
                       LIMIT 1
                   ), '')
             WHERE 1 = 1
               {electrical_filter}
            '''
        )
    if "Voltage Rating (UoM)" in existing_cols:
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Voltage Rating (UoM)" = CASE
                   WHEN TRIM(COALESCE("Voltage Rating", '')) != '' THEN 'VLT'
                   ELSE ''
               END
             WHERE COALESCE("Voltage Rating (UoM)", '') != CASE
                   WHEN TRIM(COALESCE("Voltage Rating", '')) != '' THEN 'VLT'
                   ELSE ''
               END
            '''
        )
    if "Amperage Rating" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Amperage Rating" TEXT')
    if "Amperage Rating (UoM)" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Amperage Rating (UoM)" TEXT')
        existing_cols.add("Amperage Rating (UoM)")
    if "Fed From Amperage Rating" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Fed From Amperage Rating" TEXT')
        existing_cols.add("Fed From Amperage Rating")
    if "Fed From Amperage Rating (UoM)" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Fed From Amperage Rating (UoM)" TEXT')
        existing_cols.add("Fed From Amperage Rating (UoM)")
    if "Power Rating" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Power Rating" TEXT')
        existing_cols.add("Power Rating")
    if "Power Rating (UoM)" not in existing_cols:
        cur.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Power Rating (UoM)" TEXT')
        existing_cols.add("Power Rating (UoM)")
    if "Ampere" in existing_cols:
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Amperage Rating" = TRIM(REPLACE(REPLACE(REPLACE(UPPER(COALESCE("Ampere", '')), 'AMPS', ''), 'AMP', ''), 'A', ''))
             WHERE TRIM(COALESCE("Amperage Rating", '')) = ''
               AND TRIM(COALESCE("Ampere", '')) != ''
            '''
        )
    cur.execute(
        f'''
        UPDATE "{table_name}"
           SET "Amperage Rating (UoM)" = CASE
               WHEN TRIM(COALESCE("Amperage Rating", '')) != '' THEN 'AMP'
               ELSE ''
           END
         WHERE COALESCE("Amperage Rating (UoM)", '') != CASE
               WHEN TRIM(COALESCE("Amperage Rating", '')) != '' THEN 'AMP'
               ELSE ''
           END
        '''
    )
    if "Supply From" in existing_cols and "Building" in existing_cols and "Fed From Amperage Rating" in existing_cols:
        electrical_filter = ""
        if "Attribute" in existing_cols:
            electrical_filter = ' AND LOWER(TRIM(COALESCE(child."Attribute", \'\'))) = \'electrical\''
        
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Fed From Amperage Rating" = '',
                   "Fed From Amperage Rating (UoM)" = ''
             WHERE TRIM(COALESCE("Supply From", '')) = ''
               AND TRIM(COALESCE("Fed From Amperage Rating", '')) != ''
            '''
        )
        
        if table_exists(conn, "electrical_building_schema"):
            # Fed From Amperage is sourced from the SLD diagram table; rows whose
            # feeder tag has no active SLD entry stay blank.
            cur.execute(
                f'''
                UPDATE "{table_name}" AS child
                   SET "Fed From Amperage Rating" = COALESCE((
                       SELECT TRIM(COALESCE(parent."Amperage Rating", ''))
                       FROM "electrical_building_schema" AS parent
                       WHERE UPPER(TRIM(COALESCE(parent."Building", ''))) = UPPER(TRIM(COALESCE(child."Building", '')))
                         AND UPPER(TRIM(COALESCE(parent."Equipment ID", ''))) = UPPER(TRIM(COALESCE(child."Supply From", '')))
                         AND TRIM(COALESCE(parent."new_draw", '')) = 'TRUE'
                       LIMIT 1
                   ), '')
                 WHERE TRIM(COALESCE(child."Supply From", '')) != ''
                   {electrical_filter}
                '''
            )
    if "Fed From Amperage Rating (UoM)" in existing_cols:
        cur.execute(
            f'''
            UPDATE "{table_name}"
               SET "Fed From Amperage Rating (UoM)" = CASE
                   WHEN TRIM(COALESCE("Fed From Amperage Rating", '')) != '' THEN 'A'
                   ELSE ''
               END
             WHERE COALESCE("Fed From Amperage Rating (UoM)", '') != CASE
                   WHEN TRIM(COALESCE("Fed From Amperage Rating", '')) != '' THEN 'A'
                   ELSE ''
                END
            '''
        )
    if (
        "QR Code" in existing_cols
        and "Building" in existing_cols
        and "Power Type" in existing_cols
        and table_exists(conn, "sdi_dataset_EL")
    ):
        electrical_filter = ""
        if "Attribute" in existing_cols:
            electrical_filter = ' AND LOWER(TRIM(COALESCE(child."Attribute", \'\'))) = \'electrical\''
        cur.execute(
            f'''
            UPDATE "{table_name}" AS child
               SET "Power Type" = COALESCE((
                       SELECT TRIM(COALESCE(parent."Power Type", ''))
                       FROM "sdi_dataset_EL" AS parent
                       WHERE TRIM(COALESCE(parent."Building", '')) = TRIM(COALESCE(child."Building", ''))
                         AND UPPER(TRIM(COALESCE(parent."QR Code", ''))) = UPPER(TRIM(COALESCE(child."QR Code", '')))
                       LIMIT 1
                   ), '')
             WHERE 1 = 1
               {electrical_filter}
            '''
        )
    if (
        "QR Code" in existing_cols
        and "Building" in existing_cols
        and "Power Rating" in existing_cols
        and "Power Rating (UoM)" in existing_cols
        and table_exists(conn, "sdi_dataset_EL")
    ):
        electrical_filter = ""
        if "Attribute" in existing_cols:
            electrical_filter = ' AND LOWER(TRIM(COALESCE(child."Attribute", \'\'))) = \'electrical\''
        cur.execute(
            f'''
            UPDATE "{table_name}" AS child
               SET "Power Rating" = COALESCE((
                       SELECT TRIM(COALESCE(parent."Power Rating", ''))
                       FROM "sdi_dataset_EL" AS parent
                       WHERE TRIM(COALESCE(parent."Building", '')) = TRIM(COALESCE(child."Building", ''))
                         AND UPPER(TRIM(COALESCE(parent."QR Code", ''))) = UPPER(TRIM(COALESCE(child."QR Code", '')))
                       LIMIT 1
                   ), ''),
                   "Power Rating (UoM)" = COALESCE((
                       SELECT CASE
                           WHEN TRIM(COALESCE(parent."Power Rating", '')) != '' THEN TRIM(COALESCE(parent."Power Rating (UoM)", ''))
                           ELSE ''
                       END
                       FROM "sdi_dataset_EL" AS parent
                       WHERE TRIM(COALESCE(parent."Building", '')) = TRIM(COALESCE(child."Building", ''))
                         AND UPPER(TRIM(COALESCE(parent."QR Code", ''))) = UPPER(TRIM(COALESCE(child."QR Code", '')))
                       LIMIT 1
                   ), '')
             WHERE 1 = 1
               {electrical_filter}
            '''
        )

def _package_transfer_columns() -> str:
    return ", ".join(f'"{col}"' for col in PRINT_OUT_COLS)

def _fetch_package_transfer_values(
    conn: sqlite3.Connection,
    table_name: str,
    package_ids: Iterable[object],
) -> List[tuple]:
    ids = _unique_clean(package_ids)
    if not ids or not table_exists(conn, table_name):
        return []
    transfer_cols = _package_transfer_columns()
    rows = conn.execute(
        f'''
        SELECT {transfer_cols}
        FROM "{table_name}"
        WHERE TRIM(COALESCE(CAST("id_print_out" AS TEXT), '')) IN ({_placeholders(ids)})
        ''',
        ids,
    ).fetchall()
    return [tuple(row) for row in rows]

def build_sdi_dataset(building_code: str = None) -> pd.DataFrame:
    try:
        if not qrdb.is_postgres() and not os.path.exists(DB_PATH):
            raise FileNotFoundError(f"Database not found at: {DB_PATH}")
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
            me = _read_sql("SELECT * FROM sdi_dataset;", conn)
            el = _read_sql('SELECT * FROM "sdi_dataset_EL";', conn)

        me, el = filter_approved(me), _normalize_export_source_columns(filter_approved(el))
        
        if building_code:
            me = me[me['Building'].astype(str) == str(building_code)]
            el = el[el['Building'].astype(str) == str(building_code)]

        if "UBC Asset Tag" in el.columns and "UBC Tag" not in el.columns:
            el = el.rename(columns={"UBC Asset Tag": "UBC Tag"})
        # sdi_dataset_EL stores the General-asset serial under its JSON key
        # "Serial Number" (2026-08-07 nameplate columns); the package/master
        # column is "Serial" (renamed back to "Serial Number" only at Planon
        # export). Without this rename the EL serial is silently dropped at
        # the PRINT_OUT_COLS projection in export_to_sdi.
        if "Serial Number" in el.columns and "Serial" not in el.columns:
            el = el.rename(columns={"Serial Number": "Serial"})
        el = _normalize_equipment_type_column(_normalize_equipment_id_column(el))

        return pd.concat([me, el], ignore_index=True)
    except Exception as e:
        print(f"[ERROR] in build_sdi_dataset: {repr(e)}")
        raise

def get_codes_in_print_out_table() -> set:
    try:
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
            if not table_exists(conn, "sdi_print_out"):
                return set()
            df_exp = _read_sql('SELECT DISTINCT "QR Code" FROM sdi_print_out', conn)
        return set(df_exp["QR Code"].astype(str).str.strip().tolist())
    except Exception as e:
        print(f"[ERROR] in get_codes_in_print_out_table: Could not read from sdi_print_out table: {repr(e)}")
        return set()

def get_codes_in_archive_table() -> set:
    """Retrieves QR codes that are currently in the archive table."""
    try:
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
            if not table_exists(conn, "sdi_print_out_arch"):
                return set()
            df_exp = _read_sql('SELECT DISTINCT "QR Code" FROM sdi_print_out_arch', conn)
        return set(df_exp["QR Code"].astype(str).str.strip().tolist())
    except Exception as e:
        print(f"[ERROR] in get_codes_in_archive_table: Could not read from sdi_print_out_arch table: {repr(e)}")
        return set()

def get_all_buildings() -> list:
    try:
        unpackaged_df = build_unpackaged_dataset()
        packaged_df = build_packaged_dataset()

        building_vals = set()
        for df in (unpackaged_df, packaged_df):
            if df is None or df.empty or "Building" not in df.columns:
                continue
            vals = df["Building"].dropna().astype(str).str.strip()
            building_vals.update(v for v in vals if v)

        if not building_vals:
            return []

        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
            if not table_exists(conn, 'Buildings'):
                options = []
                for code in sorted(building_vals):
                    name = f"Building {code}" if code.isdigit() else code
                    options.append({'Code': code, 'Name': name})
                return options

            df_buildings = _read_sql('SELECT "Code", "Name" FROM "Buildings"', conn)
            df_buildings['Code'] = df_buildings['Code'].astype(str).str.strip()
            df_buildings['Name'] = df_buildings['Name'].astype(str).str.strip()

            code_to_name = dict(zip(df_buildings['Code'], df_buildings['Name']))
            name_to_code = {}
            for _, row in df_buildings.iterrows():
                if row['Name'] not in name_to_code:
                    name_to_code[row['Name']] = row['Code']

            options = []
            for val in sorted(building_vals):
                if val in code_to_name:
                    code = val
                    name = code_to_name[val]
                elif val in name_to_code:
                    code = name_to_code[val]
                    name = val
                else:
                    code = val
                    name = f"Building {val}" if val.isdigit() else val
                options.append({'Code': code, 'Name': name})

            unique = {}
            for opt in options:
                unique.setdefault(opt['Code'], opt)
            return sorted(unique.values(), key=lambda x: x['Name'])
    except Exception as e:
        error_msg = f"Could not generate building list: {repr(e)}"
        print(f"[ERROR] in get_all_buildings: {error_msg}")
        flash(f"Error: {error_msg}", "danger")
        return []

def build_unpackaged_dataset(building_code: str = None) -> pd.DataFrame:
    try:
        df = build_sdi_dataset(building_code=building_code).copy()
        if df.empty:
            return pd.DataFrame()
            
        df["QR Code"] = df["QR Code"].astype(str).str.strip()
        
        # --- FIX START: Get codes from both Active and Archive tables ---
        codes_in_print_out = get_codes_in_print_out_table()
        codes_in_archive = get_codes_in_archive_table()
        
        # Combine both sets into one exclusion list
        all_excluded_codes = codes_in_print_out.union(codes_in_archive)
        
        if all_excluded_codes:
            df = df[~df["QR Code"].isin(all_excluded_codes)].copy()
        # --- FIX END ---

        try:
            with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
                ensure_qr_codes_integrity_guards(conn)
                qr_codes_df = _build_qr_codes_lookup(conn)
                qr_process_df = _build_qr_process_lookup(conn)

                df["join_key"] = _normalize_qr_key_series(df["QR Code"])
                # MASTER_COLS supplies an empty placeholder before QR enrichment;
                # remove it so the authoritative QR_codes value does not become
                # pandas-suffixed Installation Date_x / Installation Date_y columns.
                if "Installation Date" in df.columns:
                    df = df.drop(columns=["Installation Date"])
                df = pd.merge(df, qr_codes_df, on="join_key", how="left")
                df = pd.merge(df, qr_process_df, on="join_key", how="left")
                df = _merge_qr_gps_coordinates(df, conn, overwrite_from_db=True)
            
            # Filter out rows where sdi is '1'
            df = df[df["sdi"].fillna("").astype(str).str.strip() != "1"]

            # Only New Assets feed SDI. Manual Entry (2) and Update Existing (1)
            # stop in review even when curated rows remain approved.
            df = df[pd.to_numeric(df["max_col_process"], errors="coerce").fillna(-1).astype(int) == 0]

            # Fill Space column
            df['Space'] = df['Space'].fillna('').astype(str)
            df['Floor Code'] = df['Floor Code'].fillna('').astype(str)
            
            # Fill Location column
            location_empty_mask = df['Location'].fillna('').astype(str).str.strip() == ''
            df.loc[location_empty_mask, 'Location'] = df.loc[location_empty_mask, 'Space Details'].fillna('')
            
            # Drop 'sdi' column after filtering
            df = df.drop(columns=['join_key', 'Space Details', 'sdi', 'max_col_process'])

        except Exception as e:
            print(f"[ERROR] in build_unpackaged_dataset (merging data): {repr(e)}")
            return pd.DataFrame()

        return df
    except Exception as e:
        print(f"[ERROR] in build_unpackaged_dataset (main block): {repr(e)}")
        return pd.DataFrame()

def build_packaged_dataset(building_code: str = None) -> pd.DataFrame:
    try:
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
            if not table_exists(conn, 'sdi_print_out'):
                return pd.DataFrame()
            _ensure_package_amperage_columns(conn, 'sdi_print_out')
            df = _read_sql('SELECT * FROM sdi_print_out', conn)
            
            if building_code:
                building_name = None
                if table_exists(conn, 'Buildings'):
                    df_buildings = _read_sql('SELECT "Code", "Name" FROM "Buildings"', conn)
                    df_buildings['Code'] = df_buildings['Code'].astype(str)
                    match = df_buildings[df_buildings['Code'] == str(building_code)]
                    if not match.empty:
                        building_name = match['Name'].iloc[0]

                code_mask = (df['Building'].astype(str) == str(building_code))
                if building_name:
                    name_mask = (df['Building'].astype(str) == str(building_name))
                    df = df[code_mask | name_mask].copy()
                else:
                    df = df[code_mask].copy()
            df = _merge_qr_gps_coordinates(df, conn, overwrite_from_db=True)
        
        return _normalize_export_source_columns(df)
    except Exception as e:
        print(f"[ERROR] in build_packaged_dataset: {repr(e)}")
        return pd.DataFrame()

def _check_db_writable(path: str):
    if qrdb.is_postgres():
        return  # SQLite-file writability check is N/A under PostgreSQL
    folder = os.path.dirname(path) or "."
    if not os.access(folder, os.W_OK):
        raise PermissionError(f"Folder not writable: {folder}")
    if os.path.exists(path) and not os.access(path, os.W_OK):
        raise PermissionError(f"Database file is read-only: {path}")

def _safe_filename(text: str) -> str:
    s = "" if text is None else str(text)
    s = re.sub(r'[\\/:*?"<>|]', "_", s)
    return s.strip()

def _get_building_label_for_filename(df: pd.DataFrame) -> str:
    if "Building" not in df.columns or df.empty:
        return "UnknownBuilding"
    
    uniq = [str(v).strip() for v in df["Building"].fillna("").astype(str).unique()]
    uniq = [u for u in uniq if u]

    if not uniq:
        return "UnknownBuilding"
    elif len(uniq) == 1:
        return _safe_filename(uniq[0])
    else:
        return "MULTI_Building"

def _normalize_name(text: str) -> str:
    s = "" if text is None else str(text)
    s = re.sub(r"[^0-9a-zA-Z]+", " ", s).strip().lower()
    return re.sub(r"\s+", " ", s)

def _rename_validation_output(output_dir: str, temp_path: str, desired_filename: str) -> None:
    temp_base = os.path.splitext(os.path.basename(temp_path))[0]
    desired_base = os.path.splitext(os.path.basename(desired_filename))[0]

    candidates = [
        (f"errors_{temp_base}.json", f"errors_{desired_base}.json"),
        (f"{temp_base}.json", f"{desired_base}.json"),
    ]
    for src_name, dst_name in candidates:
        src = os.path.join(output_dir, src_name)
        if os.path.exists(src):
            dst = os.path.join(output_dir, dst_name)
            try:
                os.replace(src, dst)
            except OSError:
                pass
            break

def _delete_validation_outputs_for_package(sdi_control_id: str, output_dir: str) -> int:
    if not sdi_control_id or not output_dir or not os.path.isdir(output_dir):
        return 0

    safe_id = _safe_filename(sdi_control_id)
    patterns = [
        f"SDI_Process_{safe_id}_*.json",
        f"errors_SDI_Process_{safe_id}_*.json",
    ]
    removed = 0
    for pattern in patterns:
        for path in glob.glob(os.path.join(output_dir, pattern)):
            try:
                os.remove(path)
                removed += 1
            except OSError:
                pass
    return removed

def _norm_text(value: object) -> str:
    return "" if value is None else str(value).strip()

def _norm_qr(value: object) -> str:
    return _norm_text(value).upper()

def _unique_clean(values: Iterable[object], *, upper: bool = False) -> List[str]:
    seen: set[str] = set()
    cleaned: List[str] = []
    for value in values:
        text = _norm_qr(value) if upper else _norm_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        cleaned.append(text)
    return cleaned

def _placeholders(values: Iterable[object]) -> str:
    return ",".join("?" for _ in values)

def _fetch_package_rows(conn: sqlite3.Connection, table_name: str, package_ids: Iterable[object]) -> List[Dict[str, Any]]:
    ids = _unique_clean(package_ids)
    if not ids or not table_exists(conn, table_name):
        return []
    rowid_expr = "ctid" if qrdb.is_postgres() else "rowid"   # PG has no rowid
    sql = f'''
        SELECT
            {rowid_expr} AS package_rowid,
            TRIM(COALESCE(CAST("id_print_out" AS TEXT), '')) AS id_print_out,
            TRIM(COALESCE(CAST("QR Code" AS TEXT), '')) AS qr_code,
            TRIM(COALESCE(CAST("Building" AS TEXT), '')) AS building,
            TRIM(COALESCE(CAST("print_out" AS TEXT), '')) AS print_out
        FROM "{table_name}"
        WHERE TRIM(COALESCE(CAST("id_print_out" AS TEXT), '')) IN ({_placeholders(ids)})
    '''
    cur = conn.execute(sql, ids)
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

def _find_qrs_in_package_table(conn: sqlite3.Connection, table_name: str, qr_codes: Iterable[object]) -> List[str]:
    qrs = _unique_clean(qr_codes, upper=True)
    if not qrs or not table_exists(conn, table_name):
        return []
    found: set[str] = set()
    for idx in range(0, len(qrs), 500):
        part = qrs[idx:idx + 500]
        rows = conn.execute(
            f'''
            SELECT DISTINCT UPPER(TRIM(COALESCE(CAST("QR Code" AS TEXT), ''))) AS qr
            FROM "{table_name}"
            WHERE UPPER(TRIM(COALESCE(CAST("QR Code" AS TEXT), ''))) IN ({_placeholders(part)})
            ''',
            part,
        ).fetchall()
        found.update(row[0] for row in rows if row and row[0])
    return sorted(found)

def _duplicate_qrs_in_rows(rows: List[Dict[str, Any]]) -> List[str]:
    counts: Dict[str, int] = {}
    for row in rows:
        qr = _norm_qr(row.get("qr_code"))
        if qr:
            counts[qr] = counts.get(qr, 0) + 1
    return sorted(qr for qr, count in counts.items() if count > 1)

def _rows_grouped_by_package(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        package_id = _norm_text(row.get("id_print_out"))
        if not package_id:
            continue
        grouped.setdefault(package_id, []).append(row)
    return grouped

def _qrs_from_rows(rows: List[Dict[str, Any]]) -> List[str]:
    return _unique_clean((row.get("qr_code") for row in rows), upper=True)

def _build_package_action_payload(
    *,
    action: str,
    package_id: str,
    row_count: int,
    qr_codes: Iterable[object],
    source_table: str,
    target_table: str,
    extra: Dict[str, Any] | None = None,
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "action": action,
        "package_id": package_id,
        "row_count": row_count,
        "qr_codes": _unique_clean(qr_codes, upper=True),
        "source_table": source_table,
        "target_table": target_table,
    }
    if extra:
        payload.update(extra)
    return payload

def _log_package_action(
    conn: sqlite3.Connection,
    *,
    action: str,
    package_id: str,
    op_type: str,
    source_table: str,
    target_table: str,
    row_count: int,
    qr_codes: Iterable[object],
    table_name: str | None = None,
    extra: Dict[str, Any] | None = None,
) -> None:
    payload = _build_package_action_payload(
        action=action,
        package_id=package_id,
        row_count=row_count,
        qr_codes=qr_codes,
        source_table=source_table,
        target_table=target_table,
        extra=extra,
    )
    payload_json = json.dumps(payload, sort_keys=True)
    if op_type == "DELETE":
        field_change = (payload_json, '')
    elif op_type == "INSERT":
        field_change = ("", payload_json)
    else:
        field_change = (source_table, payload_json)
    _audit_log_change(
        conn,
        qr_code=None,
        app_name="sdi_process",
        table_name=table_name or target_table or source_table or "sdi_package",
        record_pk=package_id,
        op_type=op_type,
        field_changes={"package_action": field_change},
        source="human",
        description=(
            f"{action}: package={package_id}; rows={row_count}; "
            f"source={source_table}; target={target_table}; qrs={','.join(payload['qr_codes'])}"
        ),
    )

def _set_source_approval_for_qrs(conn: sqlite3.Connection, qr_codes: Iterable[object]) -> None:
    qrs = _unique_clean(qr_codes, upper=True)
    if not qrs:
        return
    for idx in range(0, len(qrs), 500):
        part = qrs[idx:idx + 500]
        placeholders = _placeholders(part)
        for table_name in ("sdi_dataset", "sdi_dataset_EL"):
            if not table_exists(conn, table_name):
                continue
            conn.execute(
                f'''
                UPDATE "{table_name}"
                   SET "Approved" = '1'
                 WHERE UPPER(TRIM(COALESCE(CAST("QR Code" AS TEXT), ''))) IN ({placeholders})
                   AND TRIM(COALESCE(CAST("Approved" AS TEXT), '')) <> '1'
                ''',
                part,
            )
        if table_exists(conn, "QR_codes"):
            conn.execute(
                f'''
                UPDATE "QR_codes"
                   SET "Approved" = '1'
                 WHERE UPPER(TRIM(COALESCE(CAST("QR_code_ID" AS TEXT), ''))) IN ({placeholders})
                   AND TRIM(COALESCE(CAST("Approved" AS TEXT), '')) <> '1'
                ''',
                part,
            )

def _json_discipline_from_filename(filename: str) -> str:
    for marker, discipline in (("_ME_", "ME"), ("_BF_", "BF"), ("_EL_", "EL")):
        if marker in filename:
            return discipline
    return ""

def _update_review_processed_json_logs(filename_mtimes: Dict[str, float]) -> None:
    grouped: Dict[str, Dict[str, float]] = {"ME": {}, "BF": {}, "EL": {}}
    for filename, mtime in filename_mtimes.items():
        discipline = _json_discipline_from_filename(filename)
        if discipline:
            grouped[discipline][filename] = mtime

    for discipline, updates in grouped.items():
        if not updates:
            continue
        log_path = REVIEW_JSON_PROCESSED_LOGS.get(discipline)
        if not log_path:
            continue
        try:
            if os.path.exists(log_path):
                with open(log_path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                if not isinstance(payload, dict):
                    payload = {}
            else:
                payload = {}
            payload.update(updates)
            with open(log_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2, sort_keys=True)
        except Exception as exc:
            print(f"[WARN] Could not update processed JSON log {log_path}: {repr(exc)}")

def _set_review_json_approval_for_qrs(qr_codes: Iterable[object]) -> int:
    qrs = set(_unique_clean(qr_codes, upper=True))
    if not qrs or not os.path.isdir(REVIEW_JSON_DIR):
        return 0

    touched: Dict[str, float] = {}
    for filename in os.listdir(REVIEW_JSON_DIR):
        if not filename.endswith(".json"):
            continue
        qr_prefix = _norm_qr(filename.split("_", 1)[0])
        if qr_prefix not in qrs:
            continue
        path = os.path.join(REVIEW_JSON_DIR, filename)
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            structured = payload.get("structured_data")
            if not isinstance(structured, dict):
                continue
            if str(structured.get("Approved", '') or "").strip() == "True":
                continue
            structured["Approved"] = "True"
            payload["structured_data"] = structured
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=4)
            touched[filename] = os.path.getmtime(path)
        except Exception as exc:
            print(f"[WARN] Could not preserve Approved in {filename}: {repr(exc)}")

    if touched:
        _update_review_processed_json_logs(touched)
    return len(touched)

def _preflight_create_sdi_package(conn: sqlite3.Connection, qr_codes: Iterable[object], *, force_replace: bool) -> None:
    qrs = _unique_clean(qr_codes, upper=True)
    if not qrs:
        raise ValueError("No QR codes were available to create an SDI package.")
    archived_overlap = _find_qrs_in_package_table(conn, "sdi_print_out_arch", qrs)
    if archived_overlap:
        raise ValueError(
            "Cannot create an SDI package because these QR codes are already archived: "
            + ", ".join(archived_overlap[:20])
        )
    active_overlap = _find_qrs_in_package_table(conn, "sdi_print_out", qrs)
    if active_overlap and not force_replace:
        raise ValueError(
            "Cannot create an SDI package because these QR codes are already in an active package: "
            + ", ".join(active_overlap[:20])
        )

def _preflight_archive_package(conn: sqlite3.Connection, package_id: str) -> tuple[List[Dict[str, Any]], List[str]]:
    package_id = _norm_text(package_id)
    rows = _fetch_package_rows(conn, "sdi_print_out", [package_id])
    if not rows:
        raise ValueError(f"Package {package_id} was not found in active packages.")
    if any(not _norm_text(row.get("id_print_out")) for row in rows):
        raise ValueError(f"Package {package_id} has rows with a blank package ID.")

    duplicate_qrs = _duplicate_qrs_in_rows(rows)
    if duplicate_qrs:
        raise ValueError(f"Package {package_id} has duplicate QR rows: {', '.join(duplicate_qrs[:20])}")

    qrs = _qrs_from_rows(rows)
    archive_overlap = _find_qrs_in_package_table(conn, "sdi_print_out_arch", qrs)
    if archive_overlap:
        raise ValueError(
            f"Package {package_id} cannot be archived because these QRs already exist in archive: "
            + ", ".join(archive_overlap[:20])
        )

    unexported_count = sum(1 for row in rows if _norm_text(row.get("print_out")) != "1")
    if unexported_count:
        raise ValueError(
            f"Package {package_id} cannot be archived because {unexported_count} of "
            f"{len(rows)} rows have not been exported to Planon."
        )
    return rows, qrs

def _preflight_exclude_package(conn: sqlite3.Connection, package_id: str) -> tuple[List[Dict[str, Any]], List[str]]:
    package_id = _norm_text(package_id)
    rows = _fetch_package_rows(conn, "sdi_print_out", [package_id])
    if not rows:
        raise ValueError(f"Package {package_id} was not found in active packages.")
    qrs = _qrs_from_rows(rows)
    if not qrs:
        raise ValueError(f"Package {package_id} has no valid QR codes to return to Unpackaged Assets.")
    return rows, qrs

def _preflight_retrieve_packages(
    conn: sqlite3.Connection,
    package_ids: Iterable[object],
) -> tuple[List[Dict[str, Any]], List[str]]:
    ids = _unique_clean(package_ids)
    if not ids:
        raise ValueError("Please select at least one package to retrieve.")
    rows = _fetch_package_rows(conn, "sdi_print_out_arch", ids)
    found_ids = {row.get("id_print_out") for row in rows}
    missing_ids = [package_id for package_id in ids if package_id not in found_ids]
    if missing_ids:
        raise ValueError("Archived package(s) not found: " + ", ".join(missing_ids[:20]))
    if any(not _norm_text(row.get("id_print_out")) for row in rows):
        raise ValueError("One or more archived rows have a blank package ID and cannot be retrieved.")

    duplicate_qrs = _duplicate_qrs_in_rows(rows)
    if duplicate_qrs:
        raise ValueError("Selected archive rows contain duplicate QRs: " + ", ".join(duplicate_qrs[:20]))

    qrs = _qrs_from_rows(rows)
    active_overlap = _find_qrs_in_package_table(conn, "sdi_print_out", qrs)
    if active_overlap:
        raise ValueError(
            "Cannot retrieve package(s) because these QRs already exist in active packages: "
            + ", ".join(active_overlap[:20])
        )
    return rows, qrs

def get_next_sdi_package_id(conn) -> str:
    cur = conn.cursor()
    
    # 1. Initialize sequence table if not exists
    if not table_exists(conn, "sdi_sequence"):
        cur.execute("CREATE TABLE sdi_sequence (last_value INTEGER)")
        cur.execute("INSERT INTO sdi_sequence (last_value) VALUES (0)")

    # 2. Get the current tracked max value
    cur.execute("SELECT last_value FROM sdi_sequence")
    try:
        seq_val = int(cur.fetchone()[0])
    except (TypeError, ValueError):
        seq_val = 0
    
    # 3. Get the actual max value used in the data tables (Active AND Archive)
    actual_max = 0
    tables_to_check = ["sdi_print_out", "sdi_print_out_arch"]
    
    for table in tables_to_check:
        if table_exists(conn, table):
            try:
                # Find max ID that looks like 'SDI-%'
                cur.execute(f"SELECT MAX(id_print_out) FROM {table} WHERE id_print_out LIKE 'SDI-%'")
                max_id_str = cur.fetchone()[0]
                if max_id_str:
                    # Extract number part
                    parts = max_id_str.split('-')
                    if len(parts) == 2 and parts[1].isdigit():
                        num = int(parts[1])
                        if num > actual_max:
                            actual_max = num
            except Exception as e:
                print(f"Warning checking max ID in {table}: {e}")

    # 4. Synchronize: If actual data has a higher ID than the sequence table, bump the sequence
    current_highest = max(seq_val, actual_max)
    
    # 5. Increment
    new_value = current_highest + 1
    
    # 6. Update sequence table
    cur.execute("UPDATE sdi_sequence SET last_value = ?", (str(new_value),))

    return f"SDI-{new_value:05d}"

def _format_archived_package_date(raw_date: str) -> str:
    date_text = "" if raw_date is None else str(raw_date).strip()
    if not date_text:
        return "Unknown Date"

    parsed_date = pd.to_datetime(date_text, errors="coerce")
    if pd.isna(parsed_date):
        return date_text
    return parsed_date.strftime("%m/%d/%Y")

def _map_archived_attribute_to_asset_type(raw_attribute: str) -> str:
    attribute = "" if raw_attribute is None else str(raw_attribute).strip().lower()
    if not attribute:
        return ""
    if attribute == "electrical":
        return "EL"
    if attribute in {"backflowdevice", "backflow"}:
        return "BF"
    return "ME"

def _summarize_archived_asset_types(attributes: List[str]) -> str:
    ordered_types: List[str] = []
    for asset_type in ["ME", "EL", "BF"]:
        if any(_map_archived_attribute_to_asset_type(attribute) == asset_type for attribute in attributes):
            ordered_types.append(asset_type)
    return ", ".join(ordered_types) if ordered_types else "Unknown"

def get_archived_packages() -> List[Dict[str, str]]:
    """Retrieves archived packages with display metadata for the restore modal."""
    try:
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
            if not table_exists(conn, 'sdi_print_out_arch'):
                return []

            rowid_expr = "ctid" if qrdb.is_postgres() else "rowid"   # PG has no rowid
            df = _read_sql(
                f'''
                SELECT
                    {rowid_expr} AS archive_rowid,
                    "id_print_out",
                    TRIM(COALESCE(CAST("Building" AS TEXT), '')) AS building_code,
                    TRIM(COALESCE(CAST("date" AS TEXT), '')) AS package_date,
                    TRIM(COALESCE(CAST("Attribute" AS TEXT), '')) AS asset_attribute
                FROM sdi_print_out_arch
                WHERE TRIM(COALESCE(CAST("id_print_out" AS TEXT), '')) != ''
                ORDER BY "id_print_out" DESC, archive_rowid ASC
                ''',
                conn
            )

        if df.empty:
            return []

        packages: List[Dict[str, str]] = []
        grouped = df.groupby("id_print_out", sort=False)
        for package_id, group in grouped:
            building_code = next((value for value in group["building_code"].tolist() if value), '')
            formatted_date = _format_archived_package_date(
                next((value for value in group["package_date"].tolist() if value), '')
            )
            display_building_code = building_code or "Unknown"
            asset_types = _summarize_archived_asset_types(group["asset_attribute"].tolist())

            packages.append({
                "id_print_out": str(package_id).strip(),
                "building_code": display_building_code,
                "package_date": formatted_date,
                "asset_types": asset_types,
                "display_label": (
                    f"{str(package_id).strip()} - "
                    f"Building Code: {display_building_code} - "
                    f"Package Date: {formatted_date} - "
                    f"Type of Asset: {asset_types}"
                ),
            })

        return packages
    except Exception as e:
        print(f"[ERROR] in get_archived_packages: {repr(e)}")
        return []

def _resolve_building_codes_for_redirect(conn: sqlite3.Connection, building_values: List[str]) -> List[str]:
    """Return canonical building codes for redirecting after archive restore."""
    cleaned_values: List[str] = []
    for value in building_values:
        cleaned = str(value or "").strip()
        if cleaned and cleaned not in cleaned_values:
            cleaned_values.append(cleaned)

    if not cleaned_values:
        return []

    if not table_exists(conn, 'Buildings'):
        return sorted(cleaned_values)

    df_buildings = _read_sql('SELECT "Code", "Name" FROM "Buildings"', conn)
    df_buildings['Code'] = df_buildings['Code'].astype(str).str.strip()
    df_buildings['Name'] = df_buildings['Name'].astype(str).str.strip()
    code_to_name = dict(zip(df_buildings['Code'], df_buildings['Name']))
    name_to_code = {}
    for _, row in df_buildings.iterrows():
        if row['Name'] and row['Name'] not in name_to_code:
            name_to_code[row['Name']] = row['Code']

    resolved_codes = []
    for value in cleaned_values:
        if value in code_to_name:
            code = value
        elif value in name_to_code:
            code = name_to_code[value]
        else:
            code = value
        if code and code not in resolved_codes:
            resolved_codes.append(code)

    return sorted(resolved_codes)

##-------------------------------------------------------------##
## Authentication Routes Blueprint                             ##
##-------------------------------------------------------------##
auth_bp = Blueprint('auth', __name__, template_folder='template')

@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username')).first()
        if user and user.check_password(request.form.get('password')):
            login_user(user, remember=request.form.get('remember'))
            next_page = request.args.get('next')
            return redirect(next_page or url_for('main.dashboard'))
        else:
            flash('Invalid username or password.', 'danger')
            return redirect(url_for('auth.login'))
            
    return render_template('login.html', title="Login - SDI Process")

@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))

app.register_blueprint(auth_bp)

##-------------------------------------------------------------##
## Main Application Routes Blueprint                           ##
##-------------------------------------------------------------##
main_bp = Blueprint('main', __name__, template_folder='template')

@main_bp.before_request
def detect_embedded_mode():
    g.embedded = request.args.get('embedded', '').lower() == 'true'

@main_bp.route("/")
@login_required
def dashboard():
    if not has_permission(current_user, "application", "sdi_process", "viewer"):
        return access_denied_response("SDI Process Application")
    try:
        selected_building_code = (request.args.get("building_code", '') or "").strip()

        all_buildings = get_all_buildings()
        # Gate: don't compute or render any asset data until the user picks a
        # building. Avoids loading the full cross-building dataset on entry and
        # blocks building-scoped actions; archive retrieval stays global so
        # archive-only buildings can be restored.
        requires_building = not selected_building_code

        archived_packages = get_archived_packages()

        if requires_building:
            unpackaged_df = pd.DataFrame(columns=MASTER_COLS)
            packaged_df = pd.DataFrame(columns=MASTER_COLS)
        else:
            unpackaged_df = build_unpackaged_dataset(building_code=selected_building_code)
            packaged_df = build_packaged_dataset(building_code=selected_building_code)

        packaged_df = ensure_columns_and_order(packaged_df)
        unpackaged_df = ensure_columns_and_order(unpackaged_df)
        
        sdi_print_controls = []
        if not packaged_df.empty and "id_print_out" in packaged_df.columns:
            sdi_print_controls = sorted(packaged_df[packaged_df["id_print_out"].notna()]["id_print_out"].unique())

        try:
            with qrdb.get_connection(sqlite_path=DB_PATH, timeout=10) as conn:
                if table_exists(conn, 'Buildings'):
                    df_buildings = _read_sql('SELECT "Code", "Name" FROM "Buildings"', conn)
                    df_buildings['Code'] = df_buildings['Code'].astype(str)
                    
                    building_map = pd.Series(df_buildings.Name.values, index=df_buildings.Code).to_dict()

                    if not unpackaged_df.empty:
                        unpackaged_df['Building'] = unpackaged_df['Building'].map(building_map).fillna(unpackaged_df['Building'])

                    if not packaged_df.empty:
                        packaged_df['Building'] = packaged_df['Building'].map(building_map).fillna(packaged_df['Building'])
        except Exception as e:
             print(f"[ERROR] in dashboard (enriching building name): {repr(e)}")
             flash("Could not display building names correctly.", "warning")

        display_rename_map = {"id_print_out": "SDI Print Control"}
        display_columns = [display_rename_map.get(c, c) for c in MASTER_COLS]
        unpackaged_df.rename(columns=display_rename_map, inplace=True)
        packaged_df.rename(columns=display_rename_map, inplace=True)

        unpackaged_df = unpackaged_df.fillna('')
        packaged_df = packaged_df.fillna('')
        
        return render_template(
            "dashboard.html",
            title="SDI - Planon Process Management",
            columns=display_columns,
            unpackaged_rows=unpackaged_df.to_dict(orient="records"),
            packaged_rows=packaged_df.to_dict(orient="records"),
            logo_main_name=LOGO_MAIN_NAME,
            logo_fac_name=LOGO_FAC_NAME,
            all_buildings=all_buildings,
            selected_building=selected_building_code,
            requires_building=requires_building,
            sdi_print_controls=sdi_print_controls,
            archived_packages=archived_packages,
            username=current_user.username,
            is_admin=is_admin(current_user),
            can_edit=has_permission(current_user, "application", "sdi_process", "editor"),
        )
    except Exception as e:
        print(f"[FATAL ERROR] in dashboard route: {repr(e)}")
        flash("A critical error occurred while loading the dashboard. Please check the console log.", "danger")
        return render_template("dashboard.html", title="Error", columns=MASTER_COLS, unpackaged_rows=[], packaged_rows=[], all_buildings=[], requires_building=True, archived_packages=[], username=current_user.username)

@main_bp.route("/export", methods=["POST"])
@login_required
@require_permission("application", "sdi_process", "editor")
def export_to_sdi():
    building_code = request.form.get("building_code")
    force_replace = request.form.get("force_replace", "false").lower() == "true"
    active_tab_anchor = request.form.get("active_tab")

    if not building_code:
        flash("To create a pack, select only one building at time", "warning")
        return redirect(url_for("main.dashboard", _anchor=active_tab_anchor))

    try:
        df = build_unpackaged_dataset(building_code=building_code) 
        if df.empty:
            flash("No new assets are available to create an SDI package for the selected building.", "info")
            return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

        required_cols = ["Description", "Asset Group", "Attribute"]
        for col in required_cols:
            if df[col].isnull().any() or df[col].astype(str).str.strip().eq('').any():
                flash('To create a package, the fields "Description", "Asset Group" and "Attribute" must be filled in', "danger")
                return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

        if not force_replace:
            existing_codes = get_codes_in_print_out_table()
            new_codes = set(df["QR Code"].astype(str).str.strip())
            duplicate_codes = list(new_codes.intersection(existing_codes))

            if duplicate_codes:
                message = f"CONFIRM:{','.join(duplicate_codes)}"
                flash(message, "confirmation")
                return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))
        
        _check_db_writable(DB_PATH)
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=20) as conn:
            cur = conn.cursor()
            
            conn.execute(f'''CREATE TABLE IF NOT EXISTS sdi_print_out ({", ".join(f'"{col}" TEXT' for col in PRINT_OUT_COLS)})''')
            _ensure_package_amperage_columns(conn, 'sdi_print_out')
            
            existing_cols = set(qrdb.table_columns(conn, "sdi_print_out"))
            if "id_print_out" not in existing_cols:
                cur.execute('ALTER TABLE sdi_print_out ADD COLUMN "id_print_out" TEXT')

            _preflight_create_sdi_package(conn, df["QR Code"].tolist(), force_replace=force_replace)
            new_package_id = get_next_sdi_package_id(conn)
            
            now = datetime.now()
            df_print = _normalize_export_source_columns(df.copy())
            for c in PRINT_OUT_COLS:
                if c not in df_print.columns:
                    df_print[c] = ""

            df_print["id_print_out"] = new_package_id
            df_print["print_out"] = "0"
            df_print["date"] = now.strftime("%Y-%m-%d")
            df_print["time"] = now.strftime("%H:%M:%S")
            df_print = df_print.loc[:, PRINT_OUT_COLS]
            df_print = df_print.fillna("")

            if force_replace:
                codes_to_replace = df_print["QR Code"].tolist()
                if codes_to_replace:
                    placeholders = ','.join('?' for _ in codes_to_replace)
                    cur.execute(f'DELETE FROM sdi_print_out WHERE "QR Code" IN ({placeholders})', codes_to_replace)

            insert_cols = list(df_print.columns)
            cur.executemany(
                f'''
                INSERT INTO sdi_print_out ({",".join(f'"{col}"' for col in insert_cols)})
                VALUES ({",".join("?" for _ in insert_cols)})
                ''',
                df_print[insert_cols].values.tolist(),
            )
            _log_package_action(
                conn,
                action="Create SDI Package",
                package_id=new_package_id,
                op_type="INSERT",
                source_table="sdi_dataset/sdi_dataset_EL",
                target_table="sdi_print_out",
                table_name="sdi_print_out",
                row_count=len(df_print),
                qr_codes=df_print["QR Code"].tolist(),
                extra={"building_code": building_code, "force_replace": force_replace},
            )
        
        if force_replace:
            flash(f"Replaced and created SDI package {new_package_id} with {len(df_print)} rows successfully.", "success")
        else:
            flash(f"Created SDI package {new_package_id} with {len(df_print)} rows successfully.", "success")

    except Exception as e:
        print(f"[ERROR] in export_to_sdi: {repr(e)}")
        flash(f"Could not create the SDI package. {str(e)}", "danger")
    
    return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

@main_bp.route("/exclude_package", methods=["POST"])
@login_required
@require_permission("application", "sdi_process", "editor")
def exclude_package():
    sdi_control_id = request.form.get("sdi_control_id")
    building_code = request.form.get("building_code")
    active_tab_anchor = request.form.get("active_tab")

    if not sdi_control_id:
        flash("⚠️ Please select a package to exclude.", "warning")
        return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

    try:
        _check_db_writable(DB_PATH)
        package_qrs: List[str] = []
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=20) as conn:
            cur = conn.cursor()
            package_rows, package_qrs = _preflight_exclude_package(conn, sdi_control_id)
            _set_source_approval_for_qrs(conn, package_qrs)
            cur.execute('DELETE FROM sdi_print_out WHERE "id_print_out" = ?', (sdi_control_id,))
            deleted_rows = cur.rowcount
            _log_package_action(
                conn,
                action="Exclude SDI Package",
                package_id=sdi_control_id,
                op_type="DELETE",
                source_table="sdi_print_out",
                target_table="Unpackaged Assets",
                table_name="sdi_print_out",
                row_count=deleted_rows,
                qr_codes=package_qrs,
                extra={"captured_rows": len(package_rows), "building_code": building_code},
            )
            conn.commit()

        json_touched = _set_review_json_approval_for_qrs(package_qrs)
        
        if deleted_rows > 0:
            _delete_validation_outputs_for_package(sdi_control_id, JSON_OUTPUT_DIR)
            flash(
                f"Package {sdi_control_id} ({deleted_rows} assets) has been excluded and returned to Unpackaged Assets."
                + (f" Review JSON approvals updated for {json_touched} file(s)." if json_touched else ""),
                "success",
            )
        else:
            flash(f"🤔 Package {sdi_control_id} was not found or was already empty.", "info")

    except Exception as e:
        print(f"[ERROR] in exclude_package: {repr(e)}")
        flash(f"⚠️ Could not exclude the package. {str(e)}", "danger")

    return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

@main_bp.route("/move_to_archive", methods=["POST"])
@login_required
@require_permission("application", "sdi_process", "editor")
def move_to_archive():
    sdi_control_id = request.form.get("sdi_control_id")
    building_code = request.form.get("building_code")
    active_tab_anchor = request.form.get("active_tab")

    if not sdi_control_id:
        flash("⚠️ Please select a package to archive.", "warning")
        return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

    try:
        _check_db_writable(DB_PATH)
        package_qrs: List[str] = []
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=20) as conn:
            cur = conn.cursor()
            
            conn.execute(f'''CREATE TABLE IF NOT EXISTS sdi_print_out_arch ({", ".join(f'"{col}" TEXT' for col in PRINT_OUT_COLS)})''')
            _ensure_package_amperage_columns(conn, 'sdi_print_out')
            _ensure_package_amperage_columns(conn, 'sdi_print_out_arch')
            transfer_cols = _package_transfer_columns()
            package_rows, package_qrs = _preflight_archive_package(conn, sdi_control_id)
            package_values = _fetch_package_transfer_values(conn, "sdi_print_out", [sdi_control_id])
            if not package_values:
                raise ValueError(f"Package {sdi_control_id} was not found in active packages.")
            _set_source_approval_for_qrs(conn, package_qrs)

            cur.execute('DELETE FROM sdi_print_out WHERE "id_print_out" = ?', (sdi_control_id,))
            deleted_rows = cur.rowcount
            if deleted_rows != len(package_values):
                raise ValueError(
                    f"Package {sdi_control_id} changed during archive. Expected {len(package_values)} rows, "
                    f"deleted {deleted_rows}."
                )

            cur.executemany(
                f'''
                INSERT INTO sdi_print_out_arch ({transfer_cols})
                VALUES ({",".join("?" for _ in PRINT_OUT_COLS)})
                ''',
                package_values,
            )
            rows_moved = len(package_values)
            
            if rows_moved > 0:
                _log_package_action(
                    conn,
                    action="Archive SDI Package",
                    package_id=sdi_control_id,
                    op_type="UPDATE",
                    source_table="sdi_print_out",
                    target_table="sdi_print_out_arch",
                    table_name="sdi_print_out_arch",
                    row_count=rows_moved,
                    qr_codes=package_qrs,
                    extra={"captured_rows": len(package_rows), "building_code": building_code},
                )
                conn.commit()
                _set_review_json_approval_for_qrs(package_qrs)
                flash(f"✅ Package {sdi_control_id} ({rows_moved} assets) has been moved to archive.", "success")
            else:
                flash(f"🤔 Package {sdi_control_id} was not found or was already empty.", "info")

    except Exception as e:
        print(f"[ERROR] in move_to_archive: {repr(e)}")
        flash(f"⚠️ Could not archive the package. {str(e)}", "danger")

    return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

@main_bp.route("/retrieve_from_archive", methods=["POST"])
@login_required
@require_permission("application", "sdi_process", "editor")
def retrieve_from_archive():
    # Get list of selected packages (allows multiple)
    packages_to_retrieve = request.form.getlist("package_ids")
    building_code = request.form.get("building_code")
    active_tab_anchor = request.form.get("active_tab")
    redirect_building_code = (building_code or "").strip()

    if not packages_to_retrieve:
        flash("Please select at least one package to retrieve.", "warning")
        return redirect(url_for("main.dashboard", building_code=redirect_building_code, _anchor=active_tab_anchor))

    try:
        _check_db_writable(DB_PATH)
        retrieved_qrs: List[str] = []
        retrieved_buildings: List[str] = []
        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=20) as conn:
            cur = conn.cursor()
            _ensure_package_amperage_columns(conn, 'sdi_print_out')
            _ensure_package_amperage_columns(conn, 'sdi_print_out_arch')
            transfer_cols = _package_transfer_columns()
            archive_rows, retrieved_qrs = _preflight_retrieve_packages(conn, packages_to_retrieve)
            retrieved_buildings = _unique_clean((row.get("building") for row in archive_rows))
            resolved_building_codes = _resolve_building_codes_for_redirect(conn, retrieved_buildings)

            # Prepare placeholders for SQL query based on number of selected packages
            clean_package_ids = _unique_clean(packages_to_retrieve)
            placeholders = ','.join('?' for _ in clean_package_ids)
            archive_values = _fetch_package_transfer_values(conn, "sdi_print_out_arch", clean_package_ids)
            if not archive_values:
                raise ValueError("No archived rows were available for the selected package(s).")
            _set_source_approval_for_qrs(conn, retrieved_qrs)

            # Delete before insert so DB overlap triggers never see the QR in both tables.
            cur.execute(f'''
                DELETE FROM sdi_print_out_arch
                WHERE "id_print_out" IN ({placeholders})
            ''', clean_package_ids)
            deleted_rows = cur.rowcount
            if deleted_rows != len(archive_values):
                raise ValueError(
                    "Archived package rows changed during retrieve. "
                    f"Expected {len(archive_values)} rows, deleted {deleted_rows}."
                )

            cur.executemany(
                f'''
                INSERT INTO sdi_print_out ({transfer_cols})
                VALUES ({",".join("?" for _ in PRINT_OUT_COLS)})
                ''',
                archive_values,
            )
            rows_retrieved = len(archive_values)

            if rows_retrieved > 0:
                for package_id, package_rows in _rows_grouped_by_package(archive_rows).items():
                    _log_package_action(
                        conn,
                        action="Retrieve SDI Package",
                        package_id=package_id,
                        op_type="UPDATE",
                        source_table="sdi_print_out_arch",
                        target_table="sdi_print_out",
                        table_name="sdi_print_out",
                        row_count=len(package_rows),
                        qr_codes=_qrs_from_rows(package_rows),
                        extra={"building_code": building_code},
                    )
                conn.commit()

                _set_review_json_approval_for_qrs(retrieved_qrs)

                if len(resolved_building_codes) == 1:
                    redirect_building_code = resolved_building_codes[0]
                    if not active_tab_anchor:
                        active_tab_anchor = "packaged-assets-pane"
                elif len(resolved_building_codes) > 1:
                    redirect_building_code = ""
                    active_tab_anchor = None
                    flash(
                        "Retrieved packages span multiple buildings. Select a building to review the active package list.",
                        "info",
                    )

                flash(f"Successfully retrieved {len(packages_to_retrieve)} package(s) containing {rows_retrieved} assets.", "success")
            else:
                flash("No assets found in the archive for the selected packages.", "info")

    except Exception as e:
        print(f"[ERROR] in retrieve_from_archive: {repr(e)}")
        flash(f"Could not retrieve packages. {str(e)}", "danger")

    return redirect(url_for("main.dashboard", building_code=redirect_building_code, _anchor=active_tab_anchor))

def _resolve_main_asset_codes(df_to_export: pd.DataFrame):
    """Resolve the Planon-export "Main Asset" cell value for each row.

    The export's "Main Asset" column (template col AQ) must carry the
    SUST System Asset Code (e.g. ``SYS0032098``). The descriptive value
    (e.g. ``"Distribution Systems"``) lives on ``sdi_dataset`` /
    ``sdi_dataset_EL`` and is joined back per QR; the SUST code is then
    looked up by ``Property.Property code = Building`` and
    ``Description = descriptive Main Asset``.

    Routing: Attribute='electrical' rows pull from sdi_dataset_EL first,
    falling back to sdi_dataset; all other rows pull from sdi_dataset
    first.

    Returns: (series, duplicate_qrs, unresolved_qrs)
      - series indexed by ``df_to_export.index`` with the resolved Asset
        Code, or empty string for empty/no-match rows.
      - duplicate_qrs: rows whose (Building, descriptive Main Asset)
        matched more than one SUST row (caller should hard-block).
      - unresolved_qrs: rows that had a non-empty (Building, descriptive)
        but no SUST match (caller logs a warning).
    """
    if df_to_export.empty:
        return pd.Series([], dtype="object"), [], []

    qr_series = df_to_export.get("QR Code", pd.Series([], dtype="object")).fillna("").astype(str).str.strip()
    attr_series = df_to_export.get("Attribute", pd.Series([], dtype="object")).fillna("").astype(str).str.strip().str.lower()
    bld_series = df_to_export.get("Building", pd.Series([], dtype="object")).fillna("").astype(str).str.strip()
    qr_codes = qr_series.tolist()
    if not qr_codes:
        return pd.Series([""] * len(df_to_export), index=df_to_export.index, dtype="object"), [], []

    me_main = {}
    el_main = {}
    sust_matches = {}
    with qrdb.get_connection(sqlite_path=DB_PATH, timeout=15) as conn:
        placeholders = ",".join("?" * len(qr_codes))
        if table_exists(conn, "sdi_dataset"):
            for qr, ma in conn.execute(
                f'SELECT "QR Code", "Main Asset" FROM "sdi_dataset" WHERE "QR Code" IN ({placeholders})',
                qr_codes,
            ):
                me_main[(qr or "").strip()] = (ma or "").strip()
        if table_exists(conn, "sdi_dataset_EL"):
            for qr, ma in conn.execute(
                f'SELECT "QR Code", "Main Asset" FROM "sdi_dataset_EL" WHERE "QR Code" IN ({placeholders})',
                qr_codes,
            ):
                el_main[(qr or "").strip()] = (ma or "").strip()

        descriptive_by_qr = {}
        building_by_qr = {}
        for qr, attr, bld in zip(qr_codes, attr_series.tolist(), bld_series.tolist()):
            if attr == "electrical":
                descriptive_by_qr[qr] = el_main.get(qr, '') or me_main.get(qr, '')
            else:
                descriptive_by_qr[qr] = me_main.get(qr, '') or el_main.get(qr, '')
            building_by_qr[qr] = bld

        needed_pairs = {
            (building_by_qr[qr], descriptive_by_qr[qr])
            for qr in qr_codes
            if descriptive_by_qr.get(qr) and building_by_qr.get(qr)
        }
        if needed_pairs and table_exists(conn, "SUST - System List"):
            buildings = sorted({b for b, _ in needed_pairs})
            descriptions = sorted({d for _, d in needed_pairs})
            bld_ph = ",".join("?" * len(buildings))
            desc_ph = ",".join("?" * len(descriptions))
            sql = (
                f'SELECT "Property.Property code", "Description", "Asset Code" '
                f'FROM "SUST - System List" '
                f'WHERE "Property.Property code" IN ({bld_ph}) '
                f'AND "Description" IN ({desc_ph})'
            )
            for prop, desc, code in conn.execute(sql, buildings + descriptions):
                key = ((prop or "").strip(), (desc or "").strip())
                if key in needed_pairs:
                    sust_matches.setdefault(key, []).append((code or "").strip())

    duplicate_qrs = []
    duplicate_keys = {k for k, v in sust_matches.items() if len(v) > 1}
    if duplicate_keys:
        for qr in qr_codes:
            key = (building_by_qr.get(qr, ''), descriptive_by_qr.get(qr, ''))
            if key in duplicate_keys:
                duplicate_qrs.append(qr)

    resolved = {}
    unresolved_qrs = []
    for qr in qr_codes:
        desc = descriptive_by_qr.get(qr, '')
        bld = building_by_qr.get(qr, '')
        if not (desc and bld):
            resolved[qr] = ""
            continue
        codes = sust_matches.get((bld, desc), [])
        if codes:
            resolved[qr] = codes[0]
        else:
            resolved[qr] = ""
            unresolved_qrs.append(qr)

    series = qr_series.map(resolved).fillna("")
    return series, duplicate_qrs, unresolved_qrs


@main_bp.route("/export-planon", methods=["POST"])
@login_required
@require_permission("application", "sdi_process", "editor")
def export_to_planon():
    building_code = request.form.get("building_code")
    sdi_control_id = request.form.get("sdi_control_id")
    force_export = request.form.get("force_planon_export", "false").lower() == "true"
    active_tab_anchor = request.form.get("active_tab")
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.form.get("ajax") == "1"
    
    try:
        if not sdi_control_id:
            if is_ajax:
                return jsonify({"status": "error", "message": "To export, you must select a unique 'SDI Print Control' value."}), 400
            flash("To export, you must select a unique 'SDI Print Control' value.", "warning")
            return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

        df = build_packaged_dataset(building_code=building_code)
        
        df = df[df["id_print_out"] == sdi_control_id].copy()

        if df.empty:
            if is_ajax:
                return jsonify({"status": "error", "message": f"No assets found for SDI Print Control '{sdi_control_id}'."}), 400
            flash(f"No assets found for SDI Print Control '{sdi_control_id}'.", "info")
            return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

        with qrdb.get_connection(sqlite_path=DB_PATH, timeout=15) as conn:
            df_asset_group = pd.DataFrame()
            if table_exists(conn, 'Asset_Group'):
                df_asset_group = _read_sql('SELECT "Name", "Full Classification" FROM "Asset_Group"', conn)
        
        if not force_export:
            already_exported = df[df["print_out"].astype(str) == "1"]
            if not already_exported.empty:
                codes = already_exported["QR Code"].tolist()
                if is_ajax:
                    return jsonify({"status": "confirm", "codes": codes}), 200
                message = f"PLANON_CONFIRM:{','.join(codes)}"
                flash(message, "planon_confirmation")
                return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

        df_to_export = df[df["print_out"].astype(str) == "0"] if not force_export else df
        if df_to_export.empty and not force_export:
             if is_ajax:
                 return jsonify({"status": "error", "message": "All assets for this package have already been exported to Planon."}), 400
             flash("All assets for this package have already been exported to Planon.", "info")
             return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

        if not df_asset_group.empty:
            panels_mask = df_to_export['Asset Group'].str.strip().str.lower() == 'panels'
            df_to_export.loc[panels_mask, 'Asset Group'] = 'EL.21.306.4067'

            other_assets_mask = ~panels_mask
            asset_groups_to_check = df_to_export.loc[other_assets_mask, 'Asset Group'].str.strip().unique()

            if len(asset_groups_to_check) > 0:
                relevant_mappings = df_asset_group[df_asset_group['Name'].str.strip().isin(asset_groups_to_check)]
                duplicated_names = relevant_mappings[relevant_mappings['Name'].duplicated()]['Name'].unique()

                if duplicated_names.any():
                    conflicting_assets = df_to_export[df_to_export['Asset Group'].isin(duplicated_names)]
                    conflicting_qr_codes = conflicting_assets['QR Code'].tolist()
                    qr_codes_str = ", ".join(conflicting_qr_codes)
                    
                    error_message = f"The Asset Group is duplicated for QR Codes: {qr_codes_str}. This field must have a unique value."
                    if is_ajax:
                        return jsonify({"status": "error", "message": error_message}), 400
                    flash(error_message, "danger")
                    return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))

                # --- FIXED LOGIC START ---
                # Strip whitespace to ensure matches
                df_asset_group['Name'] = df_asset_group['Name'].str.strip()
                
                # Create a direct dictionary mapping (Name -> Full Classification)
                group_mapping = df_asset_group.set_index('Name')['Full Classification'].dropna().to_dict()

                # Get the current values for the relevant rows, stripped of whitespace
                current_values = df_to_export.loc[other_assets_mask, 'Asset Group'].str.strip()
                
                # Map the values using the dictionary. 
                # If a match is found, it uses the dictionary value.
                # If no match is found, .fillna keeps the original value.
                df_to_export.loc[other_assets_mask, 'Asset Group'] = (
                    current_values.map(group_mapping)
                    .fillna(df_to_export.loc[other_assets_mask, 'Asset Group'])
                )
                # --- FIXED LOGIC END ---
        
        building_label = _get_building_label_for_filename(df_to_export)
        date_str = datetime.now().strftime("%m_%d_%Y")
        
        sdi_control_ids = df_to_export["id_print_out"].dropna().unique()
        sdi_control_label = ""
        if len(sdi_control_ids) == 1:
            sdi_control_label = f"{_safe_filename(sdi_control_ids[0])}_"
        elif len(sdi_control_ids) > 1:
            sdi_control_label = "MULTI-Package_"

        output_filename = f"SDI_Process_{sdi_control_label}{date_str}_{building_label}.xlsx"

        df_to_export = _normalize_export_source_columns(df_to_export.copy())
        df2 = df_to_export.rename(columns=COLUMN_RENAME_MAP)

        if 'Fed From Equipment ID' not in df2.columns:
            df2['Fed From Equipment ID'] = ''
        df2['Fed From Equipment ID'] = df2['Fed From Equipment ID'].fillna('').astype(str).str.strip()
        fed_from_equipment_fallback = pd.Series("", index=df_to_export.index, dtype="object")
        if 'Supply From' in df_to_export.columns:
            fed_from_equipment_fallback = df_to_export['Supply From'].fillna('').astype(str).str.strip()
        needs_fed_from_equipment_fallback = df2['Fed From Equipment ID'] == ''
        df2.loc[needs_fed_from_equipment_fallback, 'Fed From Equipment ID'] = fed_from_equipment_fallback.loc[needs_fed_from_equipment_fallback]
        
        # --- Use stored Equipment ID, with legacy fallback to UBC Tag ---
        if 'Equipment ID' not in df2.columns:
            df2['Equipment ID'] = ''
        df2['Equipment ID'] = df2['Equipment ID'].fillna('').astype(str).str.strip()
        equipment_fallback = pd.Series("", index=df_to_export.index, dtype="object")
        if 'UBC Tag' in df_to_export.columns:
            equipment_fallback = df_to_export['UBC Tag'].fillna('').astype(str).str.strip()
        needs_equipment_fallback = df2['Equipment ID'] == ''
        df2.loc[needs_equipment_fallback, 'Equipment ID'] = equipment_fallback.loc[needs_equipment_fallback]

        if 'Equipment Type' not in df2.columns:
            df2['Equipment Type'] = ''
        df2['Equipment Type'] = df2['Equipment Type'].fillna('').astype(str).str.strip()
        if 'Power Type' not in df2.columns:
            df2['Power Type'] = ''
        df2['Power Type'] = df2['Power Type'].fillna('').astype(str).str.strip()

        parsed_result = df2['Equipment ID'].apply(parse_electrical_equipment_metadata)
        derived_power_type = pd.Series([res[0] for res in parsed_result], index=df2.index, dtype="object")
        needs_power_type_fallback = df2['Power Type'] == ''
        df2.loc[needs_power_type_fallback, 'Power Type'] = derived_power_type.loc[needs_power_type_fallback]
        derived_equipment_type = pd.Series([res[1] for res in parsed_result], index=df2.index, dtype="object")
        needs_type_fallback = df2['Equipment Type'] == ''
        df2.loc[needs_type_fallback, 'Equipment Type'] = derived_equipment_type.loc[needs_type_fallback]
        
        df2['Space.Floor.Property'] = df2['Property']
        for name, value in CONST_COLS.items():
            df2[name] = value

        if 'Voltage Rating' in df2.columns:
            if 'Voltage Rating (UoM)' not in df2.columns:
                df2['Voltage Rating (UoM)'] = ''
            df2['Voltage Rating (UoM)'] = df2['Voltage Rating (UoM)'].fillna('').astype(str).str.strip()
            df2['Voltage Rating'] = df2['Voltage Rating'].fillna('').astype(str).str.replace(_VOLTAGE_UNIT_SUFFIX_RE, '', regex=True).str.strip()
            condition = pd.notna(df2['Voltage Rating']) & (df2['Voltage Rating'].astype(str).str.strip() != '')
            df2.loc[condition & (df2['Voltage Rating (UoM)'] == ''), 'Voltage Rating (UoM)'] = 'VLT'
            df2.loc[~condition, 'Voltage Rating (UoM)'] = ''

        if 'Amperage Rating' in df2.columns:
            if 'Amperage Rating (UoM)' not in df2.columns:
                df2['Amperage Rating (UoM)'] = ''
            df2['Amperage Rating (UoM)'] = df2['Amperage Rating (UoM)'].fillna('').astype(str).str.strip()
            condition = pd.notna(df2['Amperage Rating']) & (df2['Amperage Rating'].astype(str).str.strip() != '')
            df2.loc[condition & (df2['Amperage Rating (UoM)'] == ''), 'Amperage Rating (UoM)'] = 'AMP'
            df2.loc[~condition, 'Amperage Rating (UoM)'] = ''

        if 'Fed From Amperage Rating' in df2.columns:
            if 'Fed From Amperage Rating (UoM)' not in df2.columns:
                df2['Fed From Amperage Rating (UoM)'] = ''
            df2['Fed From Amperage Rating (UoM)'] = df2['Fed From Amperage Rating (UoM)'].fillna('').astype(str).str.strip()
            condition = pd.notna(df2['Fed From Amperage Rating']) & (df2['Fed From Amperage Rating'].astype(str).str.strip() != '')
            df2.loc[condition, 'Fed From Amperage Rating (UoM)'] = 'A'
            df2.loc[~condition, 'Fed From Amperage Rating (UoM)'] = ''

        def format_year_to_date(year_str):
            if not year_str or pd.isna(year_str):
                return year_str
            s = str(year_str).strip()
            if s.endswith('.0'):
                s = s[:-2]
            
            if s.isdigit():
                year_val = int(s)
                full_year = None
                if len(s) == 4 and 1900 < year_val < 2100:
                    full_year = year_val
                elif len(s) == 2:
                    current_year_short = datetime.now().year % 100
                    if year_val > current_year_short:
                        full_year = 1900 + year_val
                    else:
                        full_year = 2000 + year_val
                
                if full_year:
                    return f"{full_year}-01-01"
            return year_str

        if 'Date Of Manufacture Or Construction' in df2.columns:
            df2['Date Of Manufacture Or Construction'] = df2['Date Of Manufacture Or Construction'].apply(format_year_to_date)
        
        
        # --- ATTRIBUTE SET COLUMN FILTERING (DYNAMIC) ---
        # Logic: 
        # 1. Identify all "Business Attributes" defined across ALL sets in the dictionary.
        # 2. If a column is a "Business Attribute" but not allowed for the current row, clear it.
        # 3. If a column is NOT in the dictionary at all (e.g. System/Location fields), leave it alone.
        try:
            attr_sets_path = os.path.join(BASE_DIR, 'template', 'ATTRIBUTE_SETS.py')
            # Support server path as well
            if not os.path.exists(attr_sets_path):
                attr_sets_path = '/home/developer/SDI_process/template/ATTRIBUTE_SETS.py'
            
            # Load the Attribute_Set dictionary
            Attribute_Set = {}
            if os.path.exists(attr_sets_path):
                with open(attr_sets_path, 'r') as f:
                    exec_globals = {}
                    exec(f.read(), exec_globals)
                    Attribute_Set = exec_globals.get('Attribute_Set', {})
                logging.info(f"Loaded {len(Attribute_Set)} attribute sets from configuration")
                
                # Build a master set of ALL attributes managed by the configuration
                ALL_MANAGED_ATTRIBUTES = set()
                for col_list in Attribute_Set.values():
                    ALL_MANAGED_ATTRIBUTES.update(col_list)

                # Check if 'Attribute Set' column exists (it should, based on renaming)
                if 'Attribute Set' in df2.columns:
                    # Find the index of the 'Attribute Set' column
                    # We only process columns that appear AFTER this index
                    attr_col_idx = df2.columns.get_loc('Attribute Set')
                    columns_after = df2.columns[attr_col_idx + 1:]
                    
                    debug_log = os.path.join(BASE_DIR, 'attribute_filtering_debug.log')
                    with open(debug_log, 'a') as f:
                        f.write(f"\n{'='*80}\n")
                        f.write(f"Export at {datetime.now()} (Dynamic Dictionary Filtering)\n")
                        
                        for idx, row in df2.iterrows():
                            attribute_set_name = str(row.get('Attribute Set', '')).strip()
                            
                            allowed_cols = set()
                            if attribute_set_name and attribute_set_name in Attribute_Set:
                                allowed_cols = set(Attribute_Set[attribute_set_name])
                                
                            # Iterate ONLY columns located AFTER 'Attribute Set'
                            for col in columns_after:
                                # Logic:
                                # If the column is known in the dictionary (it's a business attribute),
                                # BUT it is not allowed for this specific row -> Clear it.
                                if col in ALL_MANAGED_ATTRIBUTES and col not in allowed_cols:
                                    df2.at[idx, col] = None
                                    
                                # If col is NOT in ALL_MANAGED_ATTRIBUTES (e.g. 'Status.System name'), 
                                # we do nothing (preserve it).
            else:
                logging.warning(f"ATTRIBUTE_SETS.py not found at {attr_sets_path}")
        except Exception as e:
            logging.error(f"Error in attribute filtering: {e}")
        # --- END ATTRIBUTE SET FILTERING ---

        # --- MAIN ASSET LOOKUP (SUST - System List) ---
        # Populate template col AQ ("Main Asset") with the SUST Asset Code
        # resolved from each row's (Building, descriptive Main Asset) pair.
        # Hard-block on duplicate matches; log unresolved rows.
        try:
            main_asset_series, dup_qrs, unresolved_qrs = _resolve_main_asset_codes(df_to_export)
            if dup_qrs:
                qr_codes_str = ", ".join(dup_qrs)
                error_message = (
                    f"Multiple matches in 'SUST - System List' for the (Building, Main Asset) "
                    f"combination on QR Codes: {qr_codes_str}. Resolve before exporting."
                )
                if is_ajax:
                    return jsonify({"status": "error", "message": error_message}), 400
                flash(error_message, "danger")
                return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))
            df2["Main Asset"] = main_asset_series.values if hasattr(main_asset_series, "values") else main_asset_series
            if unresolved_qrs:
                preview = unresolved_qrs[:10]
                more = " ..." if len(unresolved_qrs) > 10 else ""
                print(
                    f"[export_to_planon] No SUST match for Main Asset on "
                    f"{len(unresolved_qrs)} row(s): {preview}{more}"
                )
        except Exception as e:
            print(f"[export_to_planon] WARNING: Main Asset SUST lookup failed: {e!r}")
            df2["Main Asset"] = ""
        # --- END MAIN ASSET LOOKUP ---

        if not os.path.exists(TEMPLATE_PATH):
            raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
        
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        header_row, start_row = 9, 10
        norm_df_cols = {_normalize_name(c): c for c in df2.columns}
        mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            norm_header = _normalize_name(header_val)
            if norm_header in norm_df_cols:
                mapping[col_idx] = norm_df_cols[norm_header]
        
        if not mapping:
            raise ValueError("No template headers matched the data columns.")
        
        df_group = df2.reset_index(drop=True)
        for r, (_, row) in enumerate(df_group.iterrows(), start=start_row):
            for col_idx, df_col in mapping.items():
                val = row.get(df_col)
                ws.cell(row=r, column=col_idx, value=(None if pd.isna(val) else val))
        
        # --- NEW LOGIC: Save to Temp File -> Validate -> Send ---
        import tempfile
        from asset_template_validation_ver00 import validate_asset_data
        
        # Create a temporary file
        # We delete=False because we need to read it back and also pass the path to the validator
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        
        try:
            # 1. Save Excel to temp path
            wb.save(temp_path)
            
            # 2. Run Validation
            json_output_dir = JSON_OUTPUT_DIR

            print(f"Running validation on generated export. Output dir: {json_output_dir}")
            summary = None
            try:
                result = validate_asset_data(
                    temp_path,
                    output_dir=json_output_dir,
                    override_filename=output_filename,
                    return_summary=True
                )
                if isinstance(result, tuple):
                    _, summary = result
            except TypeError as exc:
                if "override_filename" in str(exc) or "return_summary" in str(exc):
                    result = validate_asset_data(temp_path, output_dir=json_output_dir)
                    if isinstance(result, tuple):
                        _, summary = result
                    _rename_validation_output(json_output_dir, temp_path, output_filename)
                else:
                    raise
            
            base_name = os.path.splitext(os.path.basename(output_filename))[0]
            expected_paths = [
                os.path.join(json_output_dir, f"{base_name}.json"),
                os.path.join(json_output_dir, f"errors_{base_name}.json"),
            ]
            if not any(os.path.exists(p) for p in expected_paths):
                print(f"[WARN] Validation JSON not found in output dir: {json_output_dir}")
            
            if summary and not is_ajax:
                summary_message = (
                    "Validation summary - "
                    f"Total rows: {summary.get('total_rows', 0)}; "
                    f"Correct: {summary.get('total_correct', 0)}; "
                    f"Wrong: {summary.get('total_wrong', 0)}; "
                    f"% Correct: {summary.get('pct_correct', 0)}%; "
                    f"% Wrong: {summary.get('pct_wrong', 0)}%."
                )
                flash(summary_message, "stats_popup")
            
            # 3. Mark as Printed in DB
            _check_db_writable(DB_PATH)
            with qrdb.get_connection(sqlite_path=DB_PATH, timeout=15) as conn:
                cur = conn.cursor()
                codes_to_update = df_to_export["QR Code"].tolist()
                if codes_to_update:
                    placeholders = ','.join('?' for _ in codes_to_update)
                    cur.execute(f'UPDATE sdi_print_out SET print_out = \'1\' WHERE "QR Code" IN ({placeholders})', codes_to_update)
                    _log_package_action(
                        conn,
                        action="Export to Planon",
                        package_id=sdi_control_id,
                        op_type="UPDATE",
                        source_table="sdi_print_out",
                        target_table="Planon Export",
                        table_name="sdi_print_out",
                        row_count=len(codes_to_update),
                        qr_codes=codes_to_update,
                        extra={
                            "building_code": building_code,
                            "force_export": force_export,
                            "output_filename": output_filename,
                            "validation_summary": summary or {},
                        },
                    )
                    conn.commit()

            # 4. Prepare for Download
            # Read into buffer to send
            with open(temp_path, 'rb') as f:
                buffer = BytesIO(f.read())
            
            response = send_file(
                buffer,
                as_attachment=True,
                download_name=output_filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            if summary:
                response.headers["X-Validation-Summary"] = json.dumps(summary)
            response.headers["X-Download-Filename"] = output_filename
            return response

        finally:
            # Cleanup temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)

    except Exception as e:
        print(f"[ERROR] in export_to_planon: {repr(e)}")
        if is_ajax:
            return jsonify({"status": "error", "message": f"An unexpected error occurred: {str(e)}"}), 500
        flash(f"Error: An unexpected error occurred: {str(e)}", "danger")
        return redirect(url_for("main.dashboard", building_code=building_code, _anchor=active_tab_anchor))
@main_bp.route('/api/validation_logs', methods=['GET'])
@login_required
@require_permission("application", "sdi_process", "viewer")
def get_validation_logs():
    """API endpoint to get list of available validation log files."""
    try:
        if not os.path.exists(JSON_OUTPUT_DIR):
            return jsonify({"logs": []})
        
        logs = []
        for filename in os.listdir(JSON_OUTPUT_DIR):
            if filename.startswith('errors_') and filename.endswith('.json'):
                filepath = os.path.join(JSON_OUTPUT_DIR, filename)
                try:
                    # Get file metadata
                    stat_info = os.stat(filepath)
                    modified_time = datetime.fromtimestamp(stat_info.st_mtime)
                    
                    # Read file to get error count
                    with open(filepath, 'r', encoding='utf-8') as f:
                        errors = json.load(f)
                        error_count = len(errors) if isinstance(errors, list) else 0
                    
                    # Extract package name from filename (remove 'errors_' prefix and '.json' suffix)
                    display_name = filename[7:-5]  # Remove 'errors_' and '.json'
                    
                    logs.append({
                        'filename': filename,
                        'display_name': display_name,
                        'modified': modified_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'error_count': error_count
                    })
                except Exception as e:
                    print(f"Error processing {filename}: {e}")
                    continue
        
        # Sort by modified time (newest first)
        logs.sort(key=lambda x: x['modified'], reverse=True)
        
        return jsonify({"logs": logs})
    except Exception as e:
        print(f"[ERROR] in get_validation_logs: {repr(e)}")
        return jsonify({"error": str(e)}), 500

@main_bp.route('/api/validation_log/<filename>', methods=['GET'])
@login_required
@require_permission("application", "sdi_process", "viewer")
def get_validation_log(filename):
    """API endpoint to get contents of a specific validation log file."""
    try:
        # Sanitize filename to prevent directory traversal
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({"error": "Invalid filename"}), 400
        
        if not filename.endswith('.json'):
            return jsonify({"error": "Invalid file type"}), 400
        
        filepath = os.path.join(JSON_OUTPUT_DIR, filename)
        
        if not os.path.exists(filepath):
            return jsonify({"error": "File not found"}), 404
        
        with open(filepath, 'r', encoding='utf-8') as f:
            errors = json.load(f)
        
        # Calculate summary statistics
        if isinstance(errors, list) and len(errors) > 0:
            unique_rows = set(err.get('Row', 0) for err in errors)
            summary = {
                'total_errors': len(errors),
                'affected_rows': len(unique_rows),
                'errors': errors
            }
        else:
            summary = {
                'total_errors': 0,
                'affected_rows': 0,
                'errors': []
            }
        
        return jsonify(summary)
    except Exception as e:
        print(f"[ERROR] in get_validation_log: {repr(e)}")
        return jsonify({"error": str(e)}), 500

@main_bp.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not current_user.check_password(old_password):
            flash('Your old password was entered incorrectly. Please try again.', 'danger')
            return redirect(url_for('main.change_password'))

        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('main.change_password'))

        if len(new_password) < 8:
            flash('New password must be at least 8 characters long.', 'danger')
            return redirect(url_for('main.change_password'))

        if not re.search(r'[!@#$%^&*(),.?":{}|<>]', new_password):
            flash('New password must contain at least one special character (e.g., !@#$%).', 'danger')
            return redirect(url_for('main.change_password'))

        current_user.set_password(new_password)
        db.session.commit()

        flash('Your password has been updated successfully!', 'success')
        return redirect(url_for('main.dashboard'))
        
    return render_template('change_password.html', title="Change Password")

app.register_blueprint(main_bp)

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
with app.app_context():
    db.create_all()
    ensure_user_access_table()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8003, debug=True)
