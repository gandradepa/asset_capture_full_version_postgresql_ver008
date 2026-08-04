#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Styled XLSX export for the review dashboards (EL / ME / BF).

Reuses the same workbook styling as sld_blueprint.api_download_xlsx so the
exported file matches the on-screen palette (UBC blue header, slate borders,
AI-confidence row banding).
"""

import os
import re
from io import BytesIO
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SOURCE_LABELS = {
    "new": "New Asset",
    "update": "Up Existing",
    "manual": "Manual Entry",
}

HEADER_FILL = "FF002145"
HEADER_TEXT = "FFFFFFFF"
BORDER_COLOR = "FFE2E8F0"
TITLE_COLOR = "FF002145"
SUBTITLE_COLOR = "FF64748B"
CONF_LOW = "FFFA8282"
CONF_MID = "FFFCFC9D"
CONF_HIGH = "FF84F58F"
APPROVED_GREEN = "FF059669"
APPROVED_AMBER = "FFD97706"
SUMMARY_BAND = "FFEFF6FF"
AMP_WARNING_FILL = "FFFFC7CE"

EL_DISTRIBUTION_ASSET_GROUPS = frozenset({
    "Panels",
    "Other Service and Distribution",
    "Interior Distribution Transformers",
    "Main Transformers",
    "Motor Control Centers",
    "Enclosed Circuit Breakers",
    "Automatic Transfer Switches",
})


COLUMNS = {
    "EL": [
        ("Source", "_source", "text"),
        ("QR Code", "qr_code", "text"),
        ("Capture Date", "Capture Date", "text"),
        ("Building", "building", "text"),
        ("Space", "Space", "text"),
        ("Description", "Description", "text"),
        ("UBC Asset Tag", "UBC Asset Tag", "text"),
        ("Supply From", "Supply From", "text"),
        ("Package Number", "package_id", "text"),
        ("Avg AI Conf", "Avg_ai_conf", "percent"),
        ("Comp Score", "Comp_score", "percent"),
        ("Asset Group", "Asset Group", "text"),
        ("Manual", "_manual", "bool"),
        ("Approved", "Approved", "approved"),
        ("Attribute", "Attribute", "text"),
        ("Main Asset", "Main Asset", "text"),
        ("Amperage Rating", "_amperage_pair", "text"),
        ("Voltage Rating", "Volts", "text"),
        ("Power Rating", "_power_pair", "text"),
        ("Fed From Amp Rating", "_fed_from_amp_pair", "text"),
        ("Equipment ID", "Equipment ID", "text"),
        ("Power Type", "Power Type", "text"),
        ("Captured by", "captured_by", "text"),
        ("Date", "_capture_date_only", "text"),
        ("Hour", "_capture_hour", "text"),
    ],
    "ME": [
        ("Source", "_source", "text"),
        ("QR Code", "qr_code", "text"),
        ("Capture Date", "Capture Date", "text"),
        ("Building", "building", "text"),
        ("Space", "Location", "text"),
        ("Description", "Description", "text"),
        ("UBC Tag", "UBC Tag", "text"),
        ("Year", "Year", "text"),
        ("Asset Group", "Asset Group", "text"),
        ("Attribute", "Attribute", "text"),
        ("Main Asset", "Main Asset", "text"),
        ("Package Number", "package_id", "text"),
        ("Avg AI Conf", "Avg_ai_conf", "percent"),
        ("Comp Score", "Comp_score", "percent"),
        ("Manual", "_manual", "bool"),
        ("Approved", "Approved", "approved"),
        ("Manufacturer", "Manufacturer", "text"),
        ("Model", "Model", "text"),
        ("Serial #", "Serial Number", "text"),
        ("Technical Safety BC", "Technical Safety BC", "text"),
        ("Captured by", "captured_by", "text"),
        ("Date", "_capture_date_only", "text"),
        ("Hour", "_capture_hour", "text"),
        ("Tab", "asset_type", "text"),
    ],
    "BF": [
        ("Source", "_source", "text"),
        ("QR Code", "qr_code", "text"),
        ("Capture Date", "Capture Date", "text"),
        ("Building", "building", "text"),
        ("Space", "Location", "text"),
        ("Description", "Description", "text"),
        ("Type", "Attribute", "text"),
        ("Manufacturer", "Manufacturer", "text"),
        ("Model", "Model", "text"),
        ("Serial", "Serial Number", "text"),
        ("UBC Tag", "UBC Tag", "text"),
        ("Asset Group", "Asset Group", "text"),
        ("Attribute", "Attribute", "text"),
        ("Diameter", "Diameter", "text"),
        ("Package Number", "package_id", "text"),
        ("Avg AI Conf", "Avg_ai_conf", "percent"),
        ("Comp Score", "Comp_score", "percent"),
        ("Manual", "_manual", "bool"),
        ("Approved", "Approved", "approved"),
        ("Application", "Application", "text"),
        ("Year", "Year", "text"),
    ],
}


def _join_pair(row: dict, value_key: str, uom_key: str) -> str:
    v = str(row.get(value_key) or "").strip()
    u = str(row.get(uom_key) or "").strip()
    return f"{v} {u}".strip()


def _normalize_el_amperage_number(value) -> Optional[int]:
    match = re.search(r"\d{1,4}", str(value or "").strip())
    return int(match.group(0)) if match else None


def el_amperage_exceeds_fed_from(amperage_rating, fed_from_amperage_rating) -> bool:
    amperage = _normalize_el_amperage_number(amperage_rating)
    fed_from_amperage = _normalize_el_amperage_number(fed_from_amperage_rating)
    return (
        amperage is not None
        and fed_from_amperage is not None
        and amperage > fed_from_amperage
    )


def has_el_amperage_warning(asset_group, amperage_rating, fed_from_amperage_rating) -> bool:
    if str(asset_group or "").strip() not in EL_DISTRIBUTION_ASSET_GROUPS:
        return False
    return el_amperage_exceeds_fed_from(amperage_rating, fed_from_amperage_rating)


def _enrich_row(process: str, tab: str, row: dict, meta: dict) -> dict:
    """Add computed fields (_source, _manual, paired ratings, captured date/hour)."""
    qr = str(row.get("qr_code") or "").strip()
    info = meta.get(qr) or {}

    row["_source"] = SOURCE_LABELS.get(tab, tab)
    row["_manual"] = bool(row.get("ExcludeSDI") or tab == "manual")
    row["_capture_date_only"] = info.get("date", "")
    row["_capture_hour"] = info.get("hour", "")
    # Prefer the per-QR_code_assets username over captured_by_map when available.
    if info.get("user") and not row.get("captured_by"):
        row["captured_by"] = info["user"]

    if process == "EL":
        row["_amperage_pair"] = _join_pair(row, "Amperage Rating", "Amperage Rating (UoM)")
        row["_voltage_pair"] = _join_pair(row, "Voltage Rating", "Voltage Rating (UoM)")
        row["_power_pair"] = _join_pair(row, "Power Rating", "Power Rating (UoM)")
        row["_fed_from_amp_pair"] = _join_pair(
            row, "Fed From Amperage Rating", "Fed From Amperage Rating (UoM)"
        )
        amperage_value = row.get("Amperage Rating") or row.get("Ampere")
        row["_amp_rating_warning"] = has_el_amperage_warning(
            row.get("Asset Group"), amperage_value, row.get("Fed From Amperage Rating")
        )

    return row


def _conf_band_fill(conf_value) -> Optional[PatternFill]:
    try:
        c = float(conf_value)
    except (TypeError, ValueError):
        return None
    if c < 60.0:
        color = CONF_LOW
    elif c < 85.0:
        color = CONF_MID
    else:
        color = CONF_HIGH
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _approved_true(value) -> bool:
    return str(value or "").strip().lower() in {"true", "1", "yes"}


def _embed_corner_logo(ws, logo_path: Optional[str], anchor: str = "A1") -> bool:
    """Drop the UBC Facilities logo in the corner of the sheet. Returns True
    when embedded, False when skipped (missing path, missing Pillow, IO
    error). The logo is decorative — never fail the export over it."""
    if not logo_path:
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage  # needs Pillow
        if not os.path.isfile(logo_path):
            return False
        img = XLImage(logo_path)
        img.width = 70
        img.height = 60
        ws.add_image(img, anchor)
        return True
    except Exception as e:
        print(f"[WARN] excel_export logo embedding skipped: {e}")
        return False


def build_workbook(
    process: str,
    tab: str,
    building: str,
    rows: list[dict],
    meta: Optional[dict] = None,
    process_title: str = "",
    logo_path: Optional[str] = None,
) -> bytes:
    if process not in COLUMNS:
        raise ValueError(f"Unknown process: {process!r}")
    meta = meta or {}
    columns = COLUMNS[process]

    enriched = [_enrich_row(process, tab, dict(r), meta) for r in rows]

    wb = Workbook()
    ws = wb.active
    source_label = SOURCE_LABELS.get(tab, tab.title())
    ws.title = source_label[:31]
    ws.sheet_view.showGridLines = False

    logo_embedded = _embed_corner_logo(ws, logo_path, anchor="A1")
    title_col = 2 if logo_embedded else 1
    if logo_embedded:
        ws.column_dimensions["A"].width = 11

    process_display = process_title or process
    title_cell = ws.cell(row=1, column=title_col,
                         value=f"UBC Facilities — {process_display} Review Export")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=TITLE_COLOR)
    title_cell.alignment = Alignment(vertical="center", horizontal="left")

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    sub_text = (
        f"Building: {building or '(all)'}   |   "
        f"Source: {source_label}   |   "
        f"Generated {generated_at}   |   "
        f"{len(enriched)} rows"
    )
    sub_cell = ws.cell(row=2, column=title_col, value=sub_text)
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color=SUBTITLE_COLOR)
    sub_cell.alignment = Alignment(vertical="center", horizontal="left")

    ws.row_dimensions[1].height = 36 if logo_embedded else 22
    ws.row_dimensions[2].height = 26 if logo_embedded else 18
    ws.row_dimensions[3].height = 6

    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_border = Border(
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
    )

    HEADER_ROW = 4
    for col_idx, (header, _key, _kind) in enumerate(columns, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border
        ws.column_dimensions[c.column_letter].width = max(12, min(28, len(header) + 4))
    ws.row_dimensions[HEADER_ROW].height = 26

    body_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    body_center = Alignment(horizontal="center", vertical="center")
    bold_font = Font(name="Calibri", size=11, bold=True)
    approved_font_true = Font(name="Calibri", size=11, bold=True, color=APPROVED_GREEN)
    approved_font_false = Font(name="Calibri", size=11, bold=True, color=APPROVED_AMBER)
    amp_warning_fill = PatternFill(
        start_color=AMP_WARNING_FILL, end_color=AMP_WARNING_FILL, fill_type="solid"
    )

    for offset, row in enumerate(enriched):
        r = HEADER_ROW + 1 + offset

        for col_idx, (_header, key, kind) in enumerate(columns, start=1):
            raw_val = row.get(key, "")
            cell = ws.cell(row=r, column=col_idx)
            cell.border = cell_border
            cell.alignment = body_left

            if kind == "percent":
                try:
                    val = float(raw_val) / 100.0
                    cell.value = val
                    cell.number_format = "0.00%"
                    cell.alignment = body_center
                except (TypeError, ValueError):
                    cell.value = ""
                    cell.alignment = body_center
                band_fill = _conf_band_fill(raw_val)
                if band_fill is not None:
                    cell.fill = band_fill
            elif kind == "bool":
                cell.value = "Yes" if raw_val else "No"
                cell.alignment = body_center
                if raw_val:
                    cell.font = bold_font
            elif kind == "approved":
                is_true = _approved_true(raw_val)
                cell.value = "True" if is_true else "False"
                cell.alignment = body_center
                cell.font = approved_font_true if is_true else approved_font_false
            else:
                cell.value = "" if raw_val is None else str(raw_val)

            if (
                process == "EL"
                and row.get("_amp_rating_warning")
                and key == "_amperage_pair"
            ):
                cell.fill = amp_warning_fill

    _build_summary_sheet(wb, process_display, building, tab, enriched, logo_path)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(wb, process_display: str, building: str, tab: str, rows: list[dict], logo_path: Optional[str] = None):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
    band_fill = PatternFill(start_color=SUMMARY_BAND, end_color=SUMMARY_BAND, fill_type="solid")
    label_font = Font(name="Calibri", size=11, bold=True, color=TITLE_COLOR)
    value_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=14, bold=True, color=TITLE_COLOR)

    logo_embedded = _embed_corner_logo(ws, logo_path, anchor="A1")
    title_col = 2 if logo_embedded else 1
    if logo_embedded:
        ws.column_dimensions["A"].width = 11
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 22

    ws.cell(row=1, column=title_col, value=f"UBC Facilities — {process_display} Review Summary").font = title_font
    ws.cell(row=2, column=title_col, value=f"Generated {datetime.now():%Y-%m-%d %H:%M}").font = Font(
        name="Calibri", size=10, italic=True, color=SUBTITLE_COLOR
    )
    ws.column_dimensions["B" if logo_embedded else "A"].width = 32
    ws.column_dimensions["C" if logo_embedded else "B"].width = 28

    h1 = ws.cell(row=4, column=title_col, value="Metric")
    h2 = ws.cell(row=4, column=title_col + 1, value="Value")
    for c in (h1, h2):
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    approved_count = sum(1 for r in rows if _approved_true(r.get("Approved")))
    pending_count = len(rows) - approved_count
    confs = []
    for r in rows:
        try:
            confs.append(float(r.get("Avg_ai_conf")))
        except (TypeError, ValueError):
            continue
    avg_conf_value = (sum(confs) / len(confs) / 100.0) if confs else 0.0
    counts_by_source = {label: 0 for label in SOURCE_LABELS.values()}
    for r in rows:
        s = r.get("_source") or SOURCE_LABELS.get(tab, tab)
        counts_by_source[s] = counts_by_source.get(s, 0) + 1

    summary_rows = [
        ("Building", building or "(all)"),
        ("Source (active tab)", SOURCE_LABELS.get(tab, tab)),
        ("Total rows", len(rows)),
        ("Approved", approved_count),
        ("Pending", pending_count),
        ("Average AI Confidence", avg_conf_value),
    ]
    for label, count in counts_by_source.items():
        summary_rows.append((f"Count — {label}", count))

    for idx, (label, value) in enumerate(summary_rows, start=5):
        a = ws.cell(row=idx, column=title_col, value=label)
        b = ws.cell(row=idx, column=title_col + 1, value=value)
        a.font = label_font
        b.font = value_font
        if idx % 2 == 0:
            a.fill = band_fill
            b.fill = band_fill
        if label == "Average AI Confidence":
            b.number_format = "0.00%"
